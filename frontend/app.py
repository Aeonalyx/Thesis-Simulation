"""
Interactive Streamlit frontend for registrar queue simulation.

Focus:
- Variant controls (scheduler + allocator + scenario)
- Reproducibility via random seed
- Frame-by-frame playback from engine event log
- Metrics, charts, request inspection, compare mode, exports
"""

import json
import math
import os
import sys
import time as tm
from io import BytesIO
from datetime import datetime, time
from typing import Any, Dict, List, Optional

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import streamlit as st
import requests


sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from backend.engine import (  # noqa: E402
    COLLEGES,
    DOCUMENT_COMPLEXITY,
    PRIORITY_WEIGHTS,
    _soft_cap,
    PRIORITY_ROC_WEIGHTS_FULL,
    COLLEGE_PRIORITY,
    REQUESTER_PRIORITY,
    REQUESTER_GENERATION_WEIGHTS,
    REQUESTER_PRIORITY_MAX,
    DocumentRequest,
    _duration_to_schedule,
)
from backend.roc import calculate_roc_weights  # noqa: E402
from backend.criteria import score_custom_criteria  # noqa: E402

# ============================================================================
# BACKEND API CONFIGURATION
# ============================================================================
BACKEND_URL = os.getenv("BACKEND_URL", "http://localhost:5000")

DEFAULT_CONFIG = {
    "colleges": ["COE", "CED", "CASS", "CSM", "CEBA", "CCS", "CHS"],
    "document_types": list(DOCUMENT_COMPLEXITY.keys()),
    "document_complexity": DOCUMENT_COMPLEXITY,
    "college_population": {"COE": 0.2454, "CED": 0.1921, "CASS": 0.1908, "CSM": 0.1553, "CEBA": 0.0983, "CCS": 0.0787, "CHS": 0.0394},
    "allocator_types": ["college_based", "workload_based", "pooled", "quota_free"],
    "scheduler_types": ["FCFS", "WEIGHTED"],
    "scenarios": ["baseline", "staff_absence", "peak_urgency", "workload_imbalance", "peak_period"],
    "priority_weights_base": PRIORITY_WEIGHTS,
    "priority_weights_full": PRIORITY_ROC_WEIGHTS_FULL,
    "priority_criteria_catalog": [
        {"key": key, "label": key.replace("_", " ").title(), "source": "built_in"}
        for key in list(PRIORITY_WEIGHTS.keys())
    ] + [{"key": "urgency", "label": "Urgency", "source": "built_in", "optional": True}],
    "custom_criteria": [],
    "custom_request_fields": [],
    "criteria_source_fields": {
        "requester_type": "Requester type",
        "document_type": "Document type",
        "college": "College",
        "urgency": "Urgency",
    },
    "criteria_scoring_types": {
        "category_mapping": "Category mapping",
        "duration_mapping": "Duration mapping",
        "numeric_scale": "Numeric scale",
    },
    "criteria_source_options": {},
}

@st.cache_data(ttl=60)
def fetch_backend_config():
    try:
        response = requests.get(f"{BACKEND_URL}/config", timeout=5)
        if response.status_code == 200:
            return response.json()
    except requests.exceptions.RequestException:
        pass
    return DEFAULT_CONFIG

backend_config = fetch_backend_config()
COLLEGES = backend_config.get("colleges", DEFAULT_CONFIG["colleges"])
DOCUMENT_COMPLEXITY = backend_config.get("document_complexity", DEFAULT_CONFIG["document_complexity"])
COLLEGE_POPULATION_CONFIG = backend_config.get("college_population", DEFAULT_CONFIG["college_population"])
ALLOCATOR_OPTIONS = backend_config.get("allocator_types", DEFAULT_CONFIG["allocator_types"])
SCHEDULER_OPTIONS = backend_config.get("scheduler_types", DEFAULT_CONFIG["scheduler_types"])
CRITERIA_CATALOG = backend_config.get("priority_criteria_catalog", DEFAULT_CONFIG["priority_criteria_catalog"])
CUSTOM_CRITERIA = backend_config.get("custom_criteria", [])
CUSTOM_REQUEST_FIELDS = backend_config.get("custom_request_fields", [])
CRITERIA_SOURCE_FIELDS = backend_config.get("criteria_source_fields", DEFAULT_CONFIG["criteria_source_fields"])
CRITERIA_SCORING_TYPES = backend_config.get("criteria_scoring_types", DEFAULT_CONFIG["criteria_scoring_types"])
CRITERIA_SOURCE_OPTIONS = backend_config.get("criteria_source_options", DEFAULT_CONFIG["criteria_source_options"])


CRITERIA_KEYS = [
    item.get("key")
    for item in CRITERIA_CATALOG
    if item.get("key") and item.get("key") != "urgency" and not item.get("optional")
]
CRITERIA_ORDER_SIGNATURE = "|".join(CRITERIA_KEYS)
GATE_CRITERIA_KEYS = ["completeness_of_requirements", "payment_status"]
CRITERIA_LABELS = {
    "completeness_of_requirements": "Completeness of requirements",
    "submission_time": "Submission time",
    "document_type": "Document type",
    "requester_status": "Requester status",
    "college_affiliation": "College affiliation",
    "payment_status": "Payment status",
    "urgency": "Urgency",
}
CRITERIA_LABELS.update({
    item.get("key"): item.get("label", str(item.get("key")).replace("_", " ").title())
    for item in CRITERIA_CATALOG
    if item.get("key")
})
WEIGHT_MODE_OPTIONS = {
    "roc": "ROC ranking",
    "manual": "Manual weights",
}

PEAK_DOCUMENT_BOOST_WEIGHTS = {
    "Official Transcript of Records (TOR) and Transfer Credentials (TC)": 3.0,
    "Certification": 2.0,
}



def format_criterion_label(key: str) -> str:
    return CRITERIA_LABELS.get(key, key.replace("_", " ").title())


def weight_state_key(key: str) -> str:
    return f"w_{key}"


def composition_state_key(group: str, option: Any) -> str:
    return f"composition_{group}_{str(option).replace(' ', '_').replace('/', '_')}"


def custom_request_field_map() -> Dict[str, Dict[str, Any]]:
    return {
        str(field.get("key")): field
        for field in CUSTOM_REQUEST_FIELDS
        if field.get("key")
    }


def custom_request_field_options(field_key: str) -> List[str]:
    field = custom_request_field_map().get(field_key, {})
    options = []
    for option in field.get("options", []) or []:
        label = str(option.get("label", "")).strip()
        if label:
            options.append(label)
    return options


def composition_group_keys() -> List[str]:
    return ["college", "requester_type", "document_type"] + list(custom_request_field_map().keys())


def _normalized_document_mix_defaults() -> Dict[str, float]:
    """Default document mix based on the same normalized complexity idea used by priority scoring."""
    raw_values: Dict[str, float] = {}
    for doc_type, duration_value in DOCUMENT_COMPLEXITY.items():
        base_duration, _ = _duration_to_schedule(duration_value)
        complexity_days = max(base_duration.total_seconds() / 86400.0, 1e-6)
        raw_values[str(doc_type)] = 1.0 / (1.0 + complexity_days)
    total = sum(raw_values.values())
    if total <= 0:
        return {str(key): 1.0 for key in DOCUMENT_COMPLEXITY.keys()}
    return {key: value / total for key, value in raw_values.items()}


def _college_population_defaults() -> Dict[str, float]:
    return {str(key): float(COLLEGE_POPULATION_CONFIG.get(key, 0.0)) for key in COLLEGES}


def _document_equal_defaults() -> Dict[str, float]:
    return {str(key): 1.0 for key in DOCUMENT_COMPLEXITY.keys()}


def _document_defaults_for_current_mode() -> Dict[str, float]:
    base_values = _document_equal_defaults()
    if not st.session_state.get("peak_mode", False):
        return base_values
    return {
        option: float(PEAK_DOCUMENT_BOOST_WEIGHTS.get(option, 1.0))
        for option in base_values.keys()
    }


def default_composition_values(group: str) -> Dict[str, float]:
    if group == "college":
        # College demand is controlled only by the College Mix composition.
        # The default follows actual college population weights.
        return _college_population_defaults()
    if group == "requester_type":
        return {
            str(key): float(REQUESTER_GENERATION_WEIGHTS.get(key, 0.0))
            for key in REQUESTER_PRIORITY.keys()
        }
    if group == "document_type":
        # Normal mode matches backend generation default: equal likelihood.
        # Peak Period uses the backend peak-period boost design: TOR/TC = 3,
        # Certification = 2, all other document types = 1.
        return _document_defaults_for_current_mode()
    if group in custom_request_field_map():
        options = custom_request_field_options(group)
        return {str(option): 1.0 for option in options}
    return {}

def _normalize_composition_values(values: Dict[str, float]) -> Dict[str, float]:
    cleaned = {str(key): max(0.0, float(value)) for key, value in values.items()}
    total = sum(cleaned.values())
    if total <= 0:
        return cleaned
    return {key: value / total for key, value in cleaned.items()}


def composition_lock_state_key(group: str) -> str:
    return f"composition_locked_targets_{group}"


def get_composition_locked_targets(group: str) -> Dict[str, float]:
    raw_locks = st.session_state.get(composition_lock_state_key(group), {})
    if not isinstance(raw_locks, dict):
        return {}
    defaults = default_composition_values(group)
    locks: Dict[str, float] = {}
    for option, value in raw_locks.items():
        if option not in defaults:
            continue
        try:
            locks[str(option)] = max(0.0, min(float(value), 1.0))
        except Exception:
            continue
    return locks


def _basis_for_unlocked_options(
    group: str,
    defaults: Dict[str, float],
    unlocked_options: List[str],
) -> Dict[str, float]:
    current_values = {
        option: max(0.0, float(st.session_state.get(composition_state_key(group, option), defaults.get(option, 0.0))))
        for option in unlocked_options
    }
    current_total = sum(current_values.values())
    if current_total > 0:
        return current_values

    default_values = {
        option: max(0.0, float(defaults.get(option, 0.0)))
        for option in unlocked_options
    }
    default_total = sum(default_values.values())
    if default_total > 0:
        return default_values

    return {option: 1.0 for option in unlocked_options}


def apply_locked_composition_targets(group: str):
    """Apply all locked target shares and distribute the remainder across unlocked options."""
    defaults = default_composition_values(group)
    if not defaults:
        return

    locked_targets = get_composition_locked_targets(group)
    locked_total = sum(locked_targets.values())

    # Safety fallback for old/bad state: scale locked targets if they somehow exceed 100%.
    if locked_total > 1.0:
        locked_targets = {option: value / locked_total for option, value in locked_targets.items()}
        st.session_state[composition_lock_state_key(group)] = locked_targets
        locked_total = 1.0

    unlocked_options = [option for option in defaults.keys() if option not in locked_targets]
    remaining_share = max(0.0, 1.0 - locked_total)

    for option, value in locked_targets.items():
        st.session_state[composition_state_key(group, option)] = float(value)

    if not unlocked_options:
        return

    basis = _basis_for_unlocked_options(group, defaults, unlocked_options)
    basis_total = sum(max(0.0, float(value)) for value in basis.values())
    if basis_total <= 0:
        equal_share = remaining_share / max(len(unlocked_options), 1)
        for option in unlocked_options:
            st.session_state[composition_state_key(group, option)] = equal_share
        return

    for option in unlocked_options:
        st.session_state[composition_state_key(group, option)] = (
            remaining_share * max(0.0, float(basis.get(option, 0.0))) / basis_total
        )


def set_composition_share(group: str, selected_option: str, target_share: float):
    """Lock one option to an exact share while preserving other previously locked options."""
    defaults = default_composition_values(group)
    if selected_option not in defaults:
        return

    target_share = max(0.0, min(float(target_share), 1.0))
    locked_targets = get_composition_locked_targets(group)

    other_locked_total = sum(
        value for option, value in locked_targets.items()
        if option != selected_option
    )
    if other_locked_total + target_share > 1.0:
        target_share = max(0.0, 1.0 - other_locked_total)

    locked_targets[selected_option] = target_share
    st.session_state[composition_lock_state_key(group)] = locked_targets
    apply_locked_composition_targets(group)


def reset_composition_group(group: str):
    st.session_state[composition_lock_state_key(group)] = {}
    for option, value in default_composition_values(group).items():
        st.session_state[composition_state_key(group, option)] = float(value)


def composition_from_ui(group: str) -> Dict[str, float]:
    defaults = default_composition_values(group)
    return {
        key: max(0.0, float(st.session_state.get(composition_state_key(group, key), value)))
        for key, value in defaults.items()
    }


def request_composition_from_ui() -> Dict[str, Dict[str, float]]:
    return {
        group: composition_from_ui(group)
        for group in composition_group_keys()
    }


def active_criteria() -> List[str]:
    """Return the list of criteria to render in the UI based on urgency toggle."""
    keys = [key for key in CRITERIA_KEYS if key != "urgency"]
    if st.session_state.get("urgency", False) and "urgency" not in keys:
        keys = keys + ["urgency"]
    return keys


def _sanitize_selected_weight_criteria(selected: Any, available: List[str]) -> List[str]:
    if not isinstance(selected, list):
        selected = available
    selected = [key for key in selected if key in available]
    return selected or available[:]


def sync_selected_weight_criteria():
    available = active_criteria()
    selected = _sanitize_selected_weight_criteria(
        st.session_state.get("selected_weight_criteria", available),
        available,
    )
    st.session_state.selected_weight_criteria = selected


def get_selected_weight_criteria() -> List[str]:
    available = active_criteria()
    return _sanitize_selected_weight_criteria(
        st.session_state.get("selected_weight_criteria", available),
        available,
    )


def get_criteria_order() -> List[str]:
    selected = get_selected_weight_criteria()
    saved = st.session_state.get("criteria_order", selected)
    if not isinstance(saved, list):
        saved = selected
    ordered = [key for key in saved if key in selected]
    ordered.extend([key for key in selected if key not in ordered])
    st.session_state.criteria_order = ordered
    return ordered


def get_ordered_selected_criteria() -> List[str]:
    order = get_criteria_order()
    selected = get_selected_weight_criteria()
    ordered = [key for key in order if key in selected]
    ordered.extend([key for key in selected if key not in ordered])
    return ordered


def set_criteria_order(order: List[str]):
    selected = get_selected_weight_criteria()
    st.session_state.criteria_order = [key for key in order if key in selected]


def reconcile_manual_weight_state():
    for key in active_criteria():
        state_key = weight_state_key(key)
        raw_value = st.session_state.get(state_key, 0.0)
        try:
            value = float(raw_value)
        except Exception:
            value = 0.0
        if value > 1.0:
            value = value / 100.0
        st.session_state[state_key] = max(0.0, min(value, 1.0))


def manual_weights_from_ui() -> Dict[str, float]:
    selected = get_ordered_selected_criteria()
    return {
        key: max(0.0, min(float(st.session_state.get(weight_state_key(key), 0.0)), 1.0))
        for key in selected
    }


def manual_weight_total() -> float:
    return sum(manual_weights_from_ui().values())


def manual_weights_complete(tolerance: float = 0.001) -> bool:
    return abs(manual_weight_total() - 1.0) <= tolerance


def roc_weights_from_ui() -> Dict[str, float]:
    return calculate_roc_weights(get_criteria_order())

# ============================================================================
# DATA WRAPPERS (To parse API JSON back into frontend-compatible objects)
# ============================================================================
class RequestRecord(DocumentRequest):
    def __init__(self, data: dict):
        standard_keys = {
            "request_id", "college", "document_type", "urgency", "requester_type",
            "submission_time", "completeness_of_requirements", "payment_status",
            "requirements_stage", "requirements_partial_time",
            "requirements_complete_time", "payment_time", "ready_time",
            "priority_score", "assignment_time", "completion_time",
            "assigned_staff", "is_custom",
        }
        extra_fields = {
            key: value
            for key, value in data.items()
            if key not in standard_keys
        }
        super().__init__(
            request_id=str(data.get("request_id", "")),
            college=str(data.get("college", "")),
            document_type=str(data.get("document_type", "")),
            urgency=int(data.get("urgency", 5)),
            requester_type=str(data.get("requester_type", "")),
            submission_time=self._parse_time(data.get("submission_time")) or datetime.now(),
            completeness_of_requirements=float(data.get("completeness_of_requirements", 1.0)),
            payment_status=str(data.get("payment_status", "Paid")),
            requirements_stage=str(data.get("requirements_stage", "complete")),
            requirements_partial_time=self._parse_time(data.get("requirements_partial_time")),
            requirements_complete_time=self._parse_time(data.get("requirements_complete_time")),
            payment_time=self._parse_time(data.get("payment_time")),
            ready_time=self._parse_time(data.get("ready_time")),
            priority_score=float(data.get("priority_score", 0.0)),
            assignment_time=self._parse_time(data.get("assignment_time")),
            completion_time=self._parse_time(data.get("completion_time")),
            assigned_staff=str(data.get("assigned_staff")) if data.get("assigned_staff") else None,
            is_custom=bool(data.get("is_custom", False)),
            extra_fields=extra_fields,
        )

    @staticmethod
    def _parse_time(time_str):
        if not time_str: return None
        try: return datetime.fromisoformat(str(time_str))
        except Exception: return None

class StaffRecord:
    def __init__(self, data: dict):
        self.staff_id = str(data.get("staff_id", ""))
        self.name = str(data.get("name", ""))
        self.college_affiliation = str(data.get("college_affiliation", ""))
        self.quota_limit = int(data.get("quota_limit", 20))

# ============================================================================
# PAGE CONFIG
# ============================================================================

st.set_page_config(
    page_title="Thesis Simulation Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)


def apply_dashboard_theme():
    st.markdown(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;600;700;800&family=Space+Grotesk:wght@500;700&display=swap');

        :root {
            --bg-soft: #0f0d1b;
            --panel: #161626;
            --panel-2: #1d1b31;
            --ink: #f5f3ff;
            --ink-soft: #b6b0d4;
            --line: #2f2a47;
            --accent-a: #a855f7;
            --accent-b: #7c3aed;
            --accent-c: #22d3ee;
            --ok: #10b981;
        }

        [data-testid="stAppViewContainer"] {
            background:
                radial-gradient(980px 420px at 96% 0%, rgba(168, 85, 247, 0.20), transparent 60%),
                radial-gradient(860px 360px at 0% 18%, rgba(34, 211, 238, 0.13), transparent 60%),
                linear-gradient(150deg, #0b0b12 0%, #161127 45%, #22163d 100%);
        }

        .stApp, [data-testid="stSidebar"] {
            font-family: 'Plus Jakarta Sans', 'Segoe UI', sans-serif;
            color: var(--ink);
        }

        [data-testid="stAppViewContainer"] .main,
        [data-testid="stAppViewContainer"] .main * {
            color: var(--ink);
        }

        [data-testid="stHeader"] {
            background: rgba(0, 0, 0, 0);
        }

        .main .block-container {
            padding-top: 1.2rem;
            padding-bottom: 2.2rem;
        }

        h1, h2, h3 {
            font-family: 'Space Grotesk', 'Plus Jakarta Sans', sans-serif;
            letter-spacing: 0.2px;
        }

        [data-testid="stCaptionContainer"], .stCaption {
            color: var(--ink-soft) !important;
        }

        [data-testid="stAlert"] {
            background: rgba(29, 27, 49, 0.82);
            border: 1px solid var(--line);
            color: var(--ink);
        }

        [data-testid="stSidebar"] {
            background: linear-gradient(180deg, #121023 0%, #191630 100%);
            border-right: 1px solid var(--line);
        }

        [data-testid="stSidebar"] label,
        [data-testid="stSidebar"] p,
        [data-testid="stSidebar"] span,
        [data-testid="stSidebar"] .stMarkdown {
            color: var(--ink) !important;
        }

        [data-testid="stSidebar"] .stSlider,
        [data-testid="stSidebar"] .stNumberInput,
        [data-testid="stSidebar"] .stSelectbox,
        [data-testid="stSidebar"] .stTextInput,
        [data-testid="stSidebar"] .stTimeInput {
            background: transparent;
        }

        [data-testid="stMetric"] {
            background: linear-gradient(160deg, #1a1730 0%, #141326 100%);
            border: 1px solid var(--line);
            border-radius: 14px;
            padding: 0.55rem 0.8rem;
            box-shadow: 0 10px 26px rgba(7, 6, 13, 0.45);
        }

        [data-testid="stMetricLabel"] {
            color: var(--ink-soft);
            font-weight: 600;
        }

        [data-testid="stMetricValue"] {
            color: var(--ink);
            font-weight: 800;
        }

        .stButton > button {
            border-radius: 10px;
            border: 1px solid #3c325e;
            background: linear-gradient(140deg, #1f1b33 0%, #151226 100%);
            color: #f8f7ff;
            font-weight: 700;
            box-shadow: 0 8px 18px rgba(6, 4, 12, 0.38);
        }

        .stButton > button:hover {
            border-color: #8b5cf6;
            background: linear-gradient(120deg, #6d28d9 0%, #a855f7 62%, #22d3ee 100%);
            color: #ffffff;
        }

        div[data-testid="stDataFrame"] {
            border: 1px solid var(--line);
            border-radius: 12px;
            overflow: hidden;
            background: linear-gradient(180deg, #221a39 0%, #1c1731 100%);
            box-shadow: 0 12px 28px rgba(5, 3, 10, 0.45);
        }

        div[data-testid="stDataFrame"] [role="grid"],
        div[data-testid="stDataFrame"] [data-testid="stDataFrameResizable"],
        div[data-testid="stDataFrame"] table {
            background: #241c3d !important;
            color: #e4def8 !important;
        }

        div[data-testid="stDataFrame"] [role="columnheader"],
        div[data-testid="stDataFrame"] th {
            background: #35245a !important;
            color: #f2ebff !important;
            border-bottom: 1px solid #513a7a !important;
        }

        div[data-testid="stDataFrame"] [role="gridcell"],
        div[data-testid="stDataFrame"] td {
            background: rgba(44, 34, 69, 0.55) !important;
            border-bottom: 1px solid #30254a !important;
            color: #e4def8 !important;
        }

        div[data-testid="stDataFrame"] [role="row"]:hover [role="gridcell"],
        div[data-testid="stDataFrame"] tr:hover td {
            background: rgba(88, 62, 129, 0.50) !important;
        }

        .stPlotlyChart {
            border: 1px solid var(--line);
            border-radius: 14px;
            padding: 0.35rem;
            background: linear-gradient(180deg, rgba(42, 30, 67, 0.82) 0%, rgba(27, 22, 46, 0.86) 100%);
            box-shadow: 0 14px 30px rgba(6, 4, 12, 0.45);
        }

        .theme-table-wrap {
            border: 1px solid #4a3470;
            border-radius: 12px;
            overflow: auto;
            background: linear-gradient(180deg, #1f1735 0%, #171227 100%);
            box-shadow: inset 0 0 0 1px rgba(137, 93, 205, 0.10);
        }

        .theme-table {
            width: 100%;
            border-collapse: collapse;
            font-size: 0.86rem;
            color: #e7e1fa;
            min-width: 760px;
        }

        .theme-table thead th {
            position: sticky;
            top: 0;
            z-index: 2;
            background: linear-gradient(180deg, #3c2a62 0%, #2e214c 100%);
            color: #f3edff;
            font-weight: 700;
            text-align: left;
            padding: 0.52rem 0.56rem;
            border-bottom: 1px solid #5b4487;
            white-space: nowrap;
        }

        .theme-table tbody td {
            padding: 0.44rem 0.56rem;
            border-bottom: 1px solid #30264b;
            color: #ddd4f5;
            white-space: nowrap;
        }

        .theme-table tbody tr:nth-child(odd) td {
            background: rgba(36, 27, 57, 0.72);
        }

        .theme-table tbody tr:nth-child(even) td {
            background: rgba(30, 23, 48, 0.72);
        }

        .theme-table tbody tr:hover td {
            background: rgba(110, 78, 164, 0.36);
            color: #f3edff;
        }


        .theme-table tbody td:last-child,
        .theme-table thead th:last-child {
            text-align: right;
        }

        .hero-band {
            display: flex;
            justify-content: space-between;
            align-items: center;
            gap: 1rem;
            margin: 0.1rem 0 1rem;
            padding: 0.9rem 1rem;
            border-radius: 14px;
            border: 1px solid rgba(168, 85, 247, 0.5);
            background: linear-gradient(105deg, #21143a 0%, #3f1b78 48%, #5c2a9d 78%, #2a8eb3 100%);
            color: #f8fafc;
            box-shadow: 0 18px 34px rgba(14, 9, 24, 0.52);
        }

        .hero-band p {
            margin: 0;
            line-height: 1.3;
        }

        .hero-kicker {
            font-size: 0.72rem;
            letter-spacing: 0.08em;
            text-transform: uppercase;
            opacity: 0.9;
            font-weight: 700;
        }

        .hero-title {
            font-size: 1.08rem;
            font-weight: 800;
            margin-top: 0.18rem;
        }

        .hero-sub {
            font-size: 0.82rem;
            opacity: 0.92;
            margin-top: 0.24rem;
        }

        .hero-pill {
            display: inline-flex;
            align-items: center;
            border: 1px solid rgba(255, 255, 255, 0.5);
            border-radius: 999px;
            padding: 0.28rem 0.7rem;
            font-weight: 700;
            font-size: 0.8rem;
            background: rgba(255, 255, 255, 0.14);
            white-space: nowrap;
        }

        .snap-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
            gap: 0.6rem;
            margin-top: 0.8rem;
            border-top: 1px solid rgba(255, 255, 255, 0.15);
            padding-top: 0.8rem;
            width: 100%;
        }
        .snap-item {
            background: rgba(255, 255, 255, 0.08);
            border: 1px solid rgba(255, 255, 255, 0.12);
            border-radius: 8px;
            padding: 0.4rem 0.6rem;
            display: flex;
            flex-direction: column;
            gap: 0.15rem;
        }
        .snap-label {
            font-size: 0.72rem;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            color: #b6b0d4;
            font-weight: 700;
        }
        .snap-value {
            font-size: 0.88rem;
            font-weight: 600;
            color: #f5f3ff;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


st.title("📊 Registrar Queue Simulation Dashboard")
st.caption("Real-time simulation and analysis with playback, staffing, and weighted routing insights.")
apply_dashboard_theme()


# ============================================================================
# CONSTANTS + STATE HELPERS
# ============================================================================

PRESET_FILE = os.path.join(os.path.dirname(__file__), "saved_presets.json")

SCHEDULER_OPTIONS = ["FCFS", "WEIGHTED"]
SCHEDULER_LABELS = {
    "FCFS": "FCFS (default)",
    "WEIGHTED": "Weighted (priority-based)",
}
ALLOCATOR_OPTIONS = ["college_based", "workload_based", "pooled", "quota_free"]
ALLOCATOR_LABELS = {
    "college_based": "College Based",
    "workload_based": "Workload Based",
    "pooled": "Pooled",
    "quota_free": "Quota Free",
}


def humanize_option_label(value: str) -> str:
    if not isinstance(value, str):
        return str(value)
    if value in SCHEDULER_LABELS:
        return SCHEDULER_LABELS[value]
    if value in ALLOCATOR_LABELS:
        return ALLOCATOR_LABELS[value]
    if value == "Variant":
        return "Variant"
    return value.replace("_", " ").title()


def humanize_event_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    return text.replace("_", " ").title()


def format_variant_label(scheduler: str, allocator: str) -> str:
    scheduler_label = SCHEDULER_LABELS.get(str(scheduler), str(scheduler))
    allocator_label = ALLOCATOR_LABELS.get(
        str(allocator), str(allocator).replace("_", " ").title()
    )
    return f"{scheduler_label} | {allocator_label}"

SPEED_OPTIONS = {
    "0.25x": 1.20,
    "0.50x": 0.80,
    "1.00x": 0.45,
    "2.00x": 0.20,
    "4.00x": 0.08,
    "8.00x": 0.05,
    "16.00x": 0.03,
    "Smooth": 0.025,
}

WEIGHT_DEFAULT_STATE = {
    weight_state_key(key): 0.0
    for key in CRITERIA_KEYS
}

DEFAULT_STATE = {
    "scheduler_type": "FCFS",
    "allocator_type": "college_based",
    "num_staff": len(COLLEGES),
    "quota_limit": 20,
    "enable_absence": False,
    "total_requests": 100,
    "disable_generated_requests": False,
    "num_absent_staff": 0,
    "peak_mode": False,
    "work_start_time": time(8, 0),
    "work_end_time": time(17, 0),
    "seed_mode": "Auto",
    "manual_seed": 12345,
    **WEIGHT_DEFAULT_STATE,
    "playback_frame": 0,
    "playback_frame_ui": 1,
    "playback_speed": "1.00x",
    "playback_playing": False,
    "urgency": False,
    "weight_mode": "roc",
    "selected_weight_criteria": CRITERIA_KEYS[:],
    "criteria_order": CRITERIA_KEYS[:],
    "uploaded_request_dataset": None,
    "uploaded_request_dataset_name": None,
}


def initialize_state():
    # College imbalance is now handled directly through College Mix composition.
    # Remove the old session key so it cannot silently influence defaults.
    st.session_state.pop("imbalance_factor", None)
    for key, value in DEFAULT_STATE.items():
        if key not in st.session_state:
            st.session_state[key] = value
    for group in composition_group_keys():
        for option, value in default_composition_values(group).items():
            state_key = composition_state_key(group, option)
            if state_key not in st.session_state:
                st.session_state[state_key] = float(value)
    if st.session_state.get("criteria_order_signature") != CRITERIA_ORDER_SIGNATURE:
        previous_selected = st.session_state.get("selected_weight_criteria", CRITERIA_KEYS[:])
        if not isinstance(previous_selected, list):
            previous_selected = CRITERIA_KEYS[:]
        selected_set = set(previous_selected)
        ordered_selected = [key for key in CRITERIA_KEYS if key in selected_set]
        ordered_selected.extend([
            key for key in previous_selected
            if key not in ordered_selected and key in active_criteria()
        ])
        st.session_state.selected_weight_criteria = ordered_selected or CRITERIA_KEYS[:]
        st.session_state.criteria_order = st.session_state.selected_weight_criteria[:]
        st.session_state.criteria_order_signature = CRITERIA_ORDER_SIGNATURE

    if "simulation_engine" not in st.session_state:
        st.session_state.simulation_engine = None
    if "simulation_results" not in st.session_state:
        st.session_state.simulation_results = None
    if "comparison_df" not in st.session_state:
        st.session_state.comparison_df = None
    if "last_run_config" not in st.session_state:
        st.session_state.last_run_config = None


def load_presets() -> Dict[str, Dict]:
    if not os.path.exists(PRESET_FILE):
        return {}
    try:
        with open(PRESET_FILE, "r", encoding="utf-8") as handle:
            data = json.load(handle)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def save_presets(presets: Dict[str, Dict]):
    with open(PRESET_FILE, "w", encoding="utf-8") as handle:
        json.dump(presets, handle, indent=2)


def collect_ui_config() -> Dict:
    return {
        "scheduler_type": st.session_state.scheduler_type,
        "allocator_type": st.session_state.allocator_type,
        "num_staff": int(st.session_state.num_staff),
        "quota_limit": int(st.session_state.quota_limit),
        "enable_absence": bool(st.session_state.enable_absence),
        "total_requests": int(st.session_state.total_requests),
        "peak_mode": bool(st.session_state.get("peak_mode", DEFAULT_STATE["peak_mode"])),
        "urgency": bool(st.session_state.urgency),
        "num_absent_staff": int(st.session_state.num_absent_staff),
        "work_start": st.session_state.work_start_time.strftime("%H:%M"),
        "work_end": st.session_state.work_end_time.strftime("%H:%M"),
        "request_composition": request_composition_from_ui(),
        "uploaded_request_dataset_name": st.session_state.get("uploaded_request_dataset_name"),
        "weight_mode": st.session_state.get("weight_mode", "roc"),
        "selected_weight_criteria": get_selected_weight_criteria(),
        "criteria_order": get_criteria_order(),
        "weights_raw": {
            key: float(st.session_state.get(weight_state_key(key), 0.0))
            for key in active_criteria()
        },
    }


def apply_ui_config(config: Dict):
    if not isinstance(config, dict):
        return
    
    raw = config.get("weights_raw", {})
    old_keys = {"urgency", "requester_type", "waiting_time"}
    if raw and old_keys.intersection(raw.keys()):
        st.warning(
            "This preset was saved with old weight key names and will use default weights. "
            "Re-save it after running a simulation to update it."
        )

    st.session_state.scheduler_type = config.get("scheduler_type", st.session_state.scheduler_type)
    st.session_state.allocator_type = config.get("allocator_type", st.session_state.allocator_type)
    st.session_state.num_staff = int(config.get("num_staff", st.session_state.num_staff))
    st.session_state.quota_limit = int(config.get("quota_limit", st.session_state.quota_limit))

    enable_absence = config.get("enable_absence")
    if enable_absence is None:
        enable_absence = int(config.get("num_absent_staff", 0)) > 0
    st.session_state.enable_absence = bool(enable_absence)

    st.session_state.total_requests = int(config.get("total_requests", st.session_state.total_requests))
    st.session_state.peak_mode = bool(config.get("peak_mode", st.session_state.get("peak_mode", DEFAULT_STATE["peak_mode"])))
    st.session_state.urgency = bool(config.get("urgency", st.session_state.urgency))
    st.session_state.weight_mode = config.get("weight_mode", st.session_state.get("weight_mode", "roc"))
    if st.session_state.weight_mode not in WEIGHT_MODE_OPTIONS:
        st.session_state.weight_mode = "roc"
    available = active_criteria()
    selected = config.get("selected_weight_criteria", st.session_state.get("selected_weight_criteria", available))
    if not isinstance(selected, list):
        selected = available
    st.session_state.selected_weight_criteria = [key for key in selected if key in available] or available[:]
    order = config.get("criteria_order", st.session_state.get("criteria_order", st.session_state.selected_weight_criteria))
    if not isinstance(order, list):
        order = st.session_state.selected_weight_criteria
    st.session_state.criteria_order = [
        key for key in order if key in st.session_state.selected_weight_criteria
    ]
    max_absent = max(0, st.session_state.num_staff - 1)
    if st.session_state.enable_absence and max_absent > 0:
        st.session_state.num_absent_staff = min(
            max(1, int(config.get("num_absent_staff", st.session_state.num_absent_staff))),
            max_absent,
        )
    else:
        st.session_state.num_absent_staff = 0

    try:
        work_start = datetime.strptime(config.get("work_start", "08:00"), "%H:%M").time()
        work_end = datetime.strptime(config.get("work_end", "17:00"), "%H:%M").time()
        st.session_state.work_start_time = work_start
        st.session_state.work_end_time = work_end
    except Exception:
        pass

    st.session_state.seed_mode = config.get("seed_mode", st.session_state.seed_mode)
    st.session_state.manual_seed = int(config.get("manual_seed", st.session_state.manual_seed))
    composition = config.get("request_composition", {})
    if isinstance(composition, dict):
        for group in composition_group_keys():
            group_values = composition.get(group, {})
            if not isinstance(group_values, dict):
                continue
            for option, default_value in default_composition_values(group).items():
                value = group_values.get(option, default_value)
                try:
                    value = float(value)
                except Exception:
                    value = float(default_value)
                st.session_state[composition_state_key(group, option)] = max(0.0, value)

    raw = config.get("weights_raw", {})
    for key in CRITERIA_KEYS:
        state_key = weight_state_key(key)
        value = raw.get(key, st.session_state.get(state_key, 0.0))
        try:
            value = float(value)
        except Exception:
            value = 0.0
        st.session_state[state_key] = value / 100.0 if value > 1.0 else value
    if "urgency" in raw:
        value = raw.get("urgency", st.session_state.get(weight_state_key("urgency"), 0.0))
        try:
            value = float(value)
        except Exception:
            value = 0.0
        st.session_state[weight_state_key("urgency")] = value / 100.0 if value > 1.0 else value
    reconcile_manual_weight_state()



def normalized_weights_from_ui() -> Dict[str, float]:
    if st.session_state.get("weight_mode", "roc") == "manual":
        return manual_weights_from_ui()
    return roc_weights_from_ui()


def clear_run_state():
    st.session_state.simulation_engine = None
    st.session_state.simulation_results = None
    st.session_state.comparison_df = None
    st.session_state.last_run_config = None
    st.session_state.playback_frame = 0
    st.session_state.playback_frame_ui = 1
    st.session_state.playback_playing = False


def build_api_payload() -> Dict:
    weights = normalized_weights_from_ui()
    uploaded_dataset = st.session_state.get("uploaded_request_dataset")
    use_uploaded_dataset = isinstance(uploaded_dataset, list) and len(uploaded_dataset) > 0
    peak_mode = bool(st.session_state.get("peak_mode", DEFAULT_STATE["peak_mode"])) and not use_uploaded_dataset
    scenario = "peak_period" if peak_mode else "baseline"
    payload = {
        "scheduler_type": st.session_state.scheduler_type,
        "allocator_type": st.session_state.allocator_type,
        "num_staff": int(st.session_state.get("num_staff", DEFAULT_STATE["num_staff"])),
        "quota_limit": int(st.session_state.get("quota_limit", DEFAULT_STATE["quota_limit"])),
        "total_requests": len(uploaded_dataset) if use_uploaded_dataset else int(st.session_state.get("total_requests", DEFAULT_STATE["total_requests"])),
        "urgency_base": 8 if peak_mode else 5,
        "imbalance_factor": 0,
        "num_absent_staff": int(st.session_state.num_absent_staff) if st.session_state.enable_absence else 0,
        "random_seed": None,
        "request_composition": request_composition_from_ui(),
        "work_start": st.session_state.work_start_time.strftime("%H:%M"),
        "work_end": st.session_state.work_end_time.strftime("%H:%M"),
        "priority_weights": weights,
        "scenario": scenario,
        "urgency": bool(st.session_state.urgency),
        "disable_generated_requests": bool(st.session_state.get("disable_generated_requests", False)) or use_uploaded_dataset
    }
    if use_uploaded_dataset:
        payload["custom_requests"] = uploaded_dataset
        payload["uploaded_request_dataset_mode"] = True
        payload["align_custom_dates"] = False
    return payload


def run_simulation_now():
    if (
        st.session_state.scheduler_type == "WEIGHTED"
        and st.session_state.get("weight_mode", "roc") == "manual"
        and not manual_weights_complete()
    ):
        st.error(f"Manual weights must total 1.00 before running. Current total: {manual_weight_total():.2f}")
        return

    st.session_state.comparison_df = None
    st.session_state.comparison_details = None

    payload = build_api_payload()
    with st.spinner():
        try:
            response = requests.post(f"{BACKEND_URL}/simulate", json=payload, timeout=120)
            if response.status_code == 200:
                data = response.json()
                results = data.get("results", {})
                
                completed_requests = [RequestRecord(req) for req in results.get("completed_requests", [])]
                generated_requests = [RequestRecord(req) for req in results.get("generated_requests", [])]
                waiting_queue = [RequestRecord(req) for req in results.get("waiting_queue", [])]
                staff_pool = [StaffRecord(s) for s in results.get("staff_info", [])]
                
                class MockEngine:
                    pass
                mock_engine = MockEngine()
                mock_engine.completed = completed_requests
                mock_engine.generated_requests = generated_requests
                mock_engine.waiting_queue = waiting_queue
                mock_engine.staff_pool = staff_pool
                mock_engine.priority_weights = results.get("priority_weights", {})
                mock_engine.custom_criteria = results.get("custom_criteria", CUSTOM_CRITERIA)
                mock_engine.workday_minutes = 9 * 60
                mock_engine.urgency = payload.get("urgency", False)

                work_hours = results.get("work_hours", {})
                try:
                    ws = datetime.strptime(work_hours.get("start", "08:00"), "%H:%M")
                    we = datetime.strptime(work_hours.get("end", "17:00"), "%H:%M")
                    mock_engine.workday_minutes = int((we - ws).total_seconds() / 60)
                except Exception:
                    mock_engine.workday_minutes = 9 * 60
                try:
                    run_config = results.get("run_config", {})
                    sim_start_date_str = run_config.get("sim_start_date")
                    if sim_start_date_str:
                        sim_date = datetime.fromisoformat(sim_start_date_str).date()
                    else:
                        reqs = results.get("completed_requests", [])
                        if reqs and reqs[0].get("submission_time"):
                            sim_date = datetime.fromisoformat(reqs[0]["submission_time"]).date()
                        else:
                            sim_date = datetime.now().date()
                except Exception:
                    sim_date = datetime.now().date()

                try:
                    start_str = work_hours.get("start", "08:00")
                    start_time_obj = datetime.strptime(start_str, "%H:%M").time()
                    mock_engine.start_time = datetime.combine(sim_date, start_time_obj)
                except Exception:
                    mock_engine.start_time = datetime.combine(sim_date, time(8, 0))
                    
                st.session_state.simulation_engine = mock_engine
                st.session_state.simulation_results = results
                st.session_state.last_run_config = {"engine_kwargs": payload, "run_config": payload}
                st.session_state.playback_frame = 0
                st.session_state.playback_frame_ui = 1
                st.session_state.playback_playing = False
            else:
                st.error(f"Backend error: {response.text}")
        except requests.exceptions.RequestException as e:
            st.error(f"Failed to connect to backend at {BACKEND_URL}: {e}. Ensure the Flask server is running.")



def parse_event_time(value: str) -> datetime:
    try:
        return datetime.fromisoformat(value)
    except Exception:
        return datetime.now()


def format_compact_datetime(value) -> str:
    if value in (None, "", "-"):
        return "-"
    if isinstance(value, datetime):
        dt = value
    else:
        try:
            dt = datetime.fromisoformat(str(value))
        except Exception:
            return str(value)
    return dt.strftime("%b %d %H:%M")


def format_compact_day(day_value) -> str:
    if day_value is None:
        return "-"
    try:
        return day_value.strftime("%b %d, %Y")
    except Exception:
        return str(day_value)


def build_staff_college_map(staff_pool) -> Dict[str, Dict[str, str]]:
    mapping: Dict[str, Dict[str, str]] = {}
    for staff in staff_pool:
        mapping[str(staff.staff_id)] = {
            "college": str(staff.college_affiliation),
            "name": str(getattr(staff, "name", "")),
        }
    return mapping


def format_staff_label(staff_id: Optional[str], staff_map: Dict[str, Dict[str, str]]) -> str:
    if not staff_id:
        return "UNASSIGNED"
    staff_text = str(staff_id)
    if staff_text.upper() == "UNASSIGNED":
        return "UNASSIGNED"
    meta = staff_map.get(staff_text, {})
    college = str(meta.get("college", "")).strip()
    if college:
        return f"{staff_text} ({college})"
    return staff_text


CHART_COLORWAY = ["#a855f7", "#7c3aed", "#22d3ee", "#c084fc", "#38bdf8", "#f472b6"]


def apply_plot_theme(fig: go.Figure):
    labels_outside = st.session_state.get("labels_outside", True)
    if labels_outside:
        legend_cfg = dict(bgcolor="rgba(0,0,0,0)", font=dict(color="#d9d2f0"), orientation="v", y=1, x=1.02, xanchor="left")
        margins = dict(l=48, r=180, t=64, b=64)
        x_title_standoff = 12
        y_title_standoff = 12
    else:
        legend_cfg = dict(bgcolor="rgba(0,0,0,0)", font=dict(color="#d9d2f0"), orientation="h", y=1.02, x=0.5, xanchor="center")
        margins = dict(l=48, r=64, t=64, b=64)
        x_title_standoff = 4
        y_title_standoff = 4

    fig.update_layout(
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(28,22,46,0.92)",
        colorway=CHART_COLORWAY,
        font=dict(color="#ebe5ff", family="Plus Jakarta Sans, Segoe UI, sans-serif"),
        legend=legend_cfg,
        margin=margins,
        autosize=True,
    )
    fig.update_xaxes(
        showgrid=True,
        gridcolor="rgba(111,87,164,0.28)",
        zeroline=False,
        linecolor="rgba(130,105,190,0.45)",
        automargin=True,
        title_standoff=x_title_standoff,
    )
    fig.update_yaxes(
        showgrid=True,
        gridcolor="rgba(111,87,164,0.28)",
        zeroline=False,
        linecolor="rgba(130,105,190,0.45)",
        automargin=True,
        title_standoff=y_title_standoff,
    )


def render_theme_table(df: pd.DataFrame, height_px: int = 320):
    if df is None or df.empty:
        return
    safe_df = df.fillna("")
    table_html = safe_df.to_html(index=False, classes="theme-table", border=0)
    st.markdown(
        f'<div class="theme-table-wrap" style="max-height:{int(height_px)}px;">{table_html}</div>',
        unsafe_allow_html=True,
    )


def parse_request_dataset_csv(uploaded_file) -> List[Dict[str, Any]]:
    df = pd.read_csv(uploaded_file)
    if "requester_type" not in df.columns and "requester_status" in df.columns:
        df["requester_type"] = df["requester_status"]

    # Backward/forward compatibility for exported lifecycle dataset columns.
    # The newer export uses concise event-style names, while the engine expects
    # the older internal field names when replaying the dataset.
    column_aliases = {
        "requirements_partial": "requirements_partial_time",
        "requirements_complete": "requirements_complete_time",
        "payment_paid": "payment_time",
    }
    for source_col, target_col in column_aliases.items():
        if target_col not in df.columns and source_col in df.columns:
            df[target_col] = df[source_col]

    if "requirements_stage" not in df.columns:
        df["requirements_stage"] = "complete"
    if "completeness_of_requirements" not in df.columns:
        df["completeness_of_requirements"] = 1.0
    if "payment_status" not in df.columns:
        df["payment_status"] = "Paid"

    required = {"college", "document_type", "requester_type", "urgency", "submission_time"}
    missing = sorted(required - set(df.columns))
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    rows = []
    for index, row in df.iterrows():
        item = {}
        for column, value in row.items():
            if pd.isna(value):
                item[column] = None
            elif hasattr(value, "item"):
                item[column] = value.item()
            else:
                item[column] = value
        item["request_id"] = item.get("request_id") or f"UPLOAD{index + 1:04d}"
        item["urgency"] = int(item.get("urgency") or 5)
        source = str(item.get("source", "")).strip().lower()
        item["is_custom"] = bool(item.get("is_custom", False)) or source == "custom"
        rows.append(item)
    return rows

def run_variant_for_figure(scheduler_type: str, allocator_type: str) -> Optional[Dict]:
    last_run = st.session_state.get("last_run_config")
    if not last_run or "engine_kwargs" not in last_run:
        return None
    
    payload = build_api_payload()
    payload["scheduler_type"] = scheduler_type
    payload["allocator_type"] = allocator_type
    
    try:
        response = requests.post(f"{BACKEND_URL}/simulate", json=payload, timeout=120)
        if response.status_code == 200:
            return response.json().get("results", {})
    except requests.exceptions.RequestException:
        pass
    return None


def build_baseline_queue_dynamics_chart(event_log: List[Dict], variant_label: str = "") -> go.Figure:
    if not event_log:
        return go.Figure()

    sorted_log = sorted(
        event_log,
        key=lambda ev: (
            parse_event_time(str(ev.get("time", ""))),
            ev.get("sequence", 0),
        ),
    )

    active_requests = set()
    queue_sizes = []
    avg_waits = []
    time_points = []
    observed_waits = []

    for event in sorted_log:
        event_type = str(event.get("event_type", "")).upper()
        request_id = event.get("request_id")

        if event_type == "ARRIVAL" and request_id:
            active_requests.add(request_id)
        elif event_type == "ASSIGN" and request_id:
            active_requests.discard(request_id)
            wait_hours = float(event.get("queue_wait_hours", 0.0) or 0.0)
            observed_waits.append(wait_hours)

        current_time = parse_event_time(str(event.get("time", "")))
        time_points.append(current_time)
        queue_sizes.append(len(active_requests))
        avg_waits.append(round(sum(observed_waits) / len(observed_waits), 2) if observed_waits else 0.0)

    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(
        go.Scatter(
            x=time_points,
            y=queue_sizes,
            mode="lines+markers",
            name="Queue Size",
            marker=dict(size=6),
            line=dict(width=2, color="#a855f7"),
        ),
        secondary_y=False,
    )
    fig.add_trace(
        go.Scatter(
            x=time_points,
            y=avg_waits,
            mode="lines+markers",
            name="Avg Waiting Time (h)",
            marker=dict(size=6),
            line=dict(width=2, color="#22d3ee"),
        ),
        secondary_y=True,
    )

    chart_title = "Queue Dynamics and Waiting Time Trend"
    if variant_label:
        chart_title = f"{chart_title} — {variant_label}"

    fig.update_layout(
        title=chart_title,
        height=420,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
    )
    fig.update_xaxes(title_text="Simulation Time")
    fig.update_yaxes(title_text="Queue Size", secondary_y=False)
    fig.update_yaxes(title_text="Average Queue Wait (h)", secondary_y=True)
    apply_plot_theme(fig)
    return fig


def build_weighted_priority_distribution_chart(
    requests: List[Dict], selected_doc_types: List[str], variant_label: str = ""
) -> go.Figure:
    rows = []
    for req in requests:
        if not req:
            continue

        if isinstance(req, dict):
            score = float(req.get("priority_score", 0.0) or 0.0)
            assigned = bool(req.get("assignment_time"))
            doc_type = req.get("document_type", "Unknown")
        else:
            score = float(getattr(req, "priority_score", 0.0) or 0.0)
            assigned = getattr(req, "assignment_time", None) is not None
            doc_type = getattr(req, "document_type", "Unknown")

        if selected_doc_types and doc_type not in selected_doc_types:
            continue

        rows.append(
            {
                "Priority Score": score,
                "Status": "Assigned" if assigned else "Unassigned",
                "Document Type": doc_type,
            }
        )

    if not rows:
        return go.Figure()

    df = pd.DataFrame(rows)
    chart_title = "Priority Score Distribution by Request Type"
    if variant_label:
        chart_title = f"{chart_title} — {variant_label}"

    fig = px.histogram(
        df,
        x="Priority Score",
        color="Document Type",
        barmode="overlay",
        nbins=20,
        histnorm="percent",
        title=chart_title,
        labels={
            "Priority Score": "Priority Score",
            "Document Type": "Request Type",
        },
        height=420,
    )
    fig.update_traces(opacity=0.75)
    fig.update_layout(
        legend=dict(title="Request Type", orientation="h", y=1.02, x=0.5, xanchor="center")
    )
    apply_plot_theme(fig)
    return fig


def build_workload_imbalance_chart(compare_df: pd.DataFrame, title: str = "Workload Imbalance and Utilization Variance") -> go.Figure:
    if compare_df is None or compare_df.empty:
        return go.Figure()

    df = compare_df.copy()
    df["variant"] = df.apply(
        lambda row: format_variant_label(row["scheduler"], row["allocator"]),
        axis=1,
    )

    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(
        go.Bar(
            x=df["variant"],
            y=df["staff_load_std"],
            name="Staff Load Std Dev",
            marker_color="#a855f7",
            text=df["staff_load_std"],
            textposition="outside",
            customdata=df[["avg_waiting_time_hours"]],
            hovertemplate="%{x}<br>Std Dev: %{y}<br>Avg Wait: %{customdata[0]} h<extra></extra>",
        ),
        secondary_y=False,
    )
    fig.add_trace(
        go.Scatter(
            x=df["variant"],
            y=df["staff_load_cv"],
            name="Staff Load CV",
            mode="lines+markers",
            marker=dict(color="#22d3ee", size=8),
            line=dict(width=2),
        ),
        secondary_y=True,
    )
    fig.update_layout(
        title="Workload Imbalance by Variant",
        height=500,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
    )
    fig.update_xaxes(title_text="Variant")
    fig.update_yaxes(title_text="Staff Load Std Dev", secondary_y=False)
    fig.update_yaxes(title_text="Staff Load CV", secondary_y=True)
    apply_plot_theme(fig)
    return fig


def build_variant_summary_chart(compare_df: pd.DataFrame, title: str = "Summary of Variant Performance") -> go.Figure:
    if compare_df is None or compare_df.empty:
        return go.Figure()

    df = compare_df.copy()

    df["variant"] = df.apply(
        lambda row: format_variant_label(row["scheduler"], row["allocator"]),
        axis=1,
    )

    summary_df = df.melt(
        id_vars=["variant"],
        value_vars=[
            "avg_waiting_time_hours",
            "throughput_req_per_day",
            "staff_load_std",
        ],
        var_name="metric",
        value_name="value",
    )

    metric_names = {
        "avg_waiting_time_hours": "Avg Waiting Time (h)",
        "throughput_req_per_day": "Throughput (req/day)",
        "staff_load_std": "Staff Load Std Dev",
    }

    summary_df["metric"] = summary_df["metric"].map(metric_names)

    summary_df["result_no"] = summary_df.groupby("metric").cumcount() + 1

    fig = px.bar(
        summary_df,
        x="variant",
        y="value",
        color="metric",
        barmode="group",
        title=title,
        height=500,
        labels={"variant": "Variant", "value": "Metric Value", "metric": "Metric"},
        text="value",
    )

    fig.update_traces(
        textposition="outside",
        hovertemplate=(
            "<b>%{x}</b><br>"
            "Metric: %{legendgroup}<br>"
            "Value: %{y:.4f}<br>"
            "Result #: %{customdata}<extra></extra>"
        ),
        customdata=summary_df[["result_no"]],
    )

    apply_plot_theme(fig)

    return fig

def routing_events(event_log: List[Dict]) -> List[Dict]:
    """Keep only request-routing decisions for request-by-request playback."""
    decision_types = {"ASSIGN", "WAITING"}
    return [event for event in event_log if event.get("event_type") in decision_types]


def playback_state(decisions: List[Dict], step: int) -> Dict:
    if not decisions:
        return {
            "current_event": None,
            "assignments": [],
            "waiting": [],
            "processed_count": 0,
            "assigned_count": 0,
            "waiting_count": 0,
            "staff_flow": {},
        }

    step = max(0, min(step, len(decisions) - 1))
    chunk = decisions[: step + 1]

    assignments = []
    waiting = []
    staff_flow: Dict[str, List[str]] = {}

    for item in chunk:
        kind = item.get("event_type")
        if kind == "ASSIGN":
            assignments.append(
                {
                    "Time": item.get("time"),
                    "Request": item.get("request_id"),
                    "College": item.get("college"),
                    "Staff": item.get("staff_id"),
                    "Priority Score": item.get("priority_score", 0.0),
                    "Queue Wait (h)": item.get("queue_wait_hours", "-"),
                    "Mode": humanize_event_text(item.get("details", "")),
                }
            )
            staff_key = item.get("staff_id") or "UNASSIGNED"
            staff_flow.setdefault(staff_key, []).append(item.get("request_id"))
        elif kind == "WAITING":
            waiting.append(
                {
                    "Time": item.get("time"),
                    "Request": item.get("request_id"),
                    "College": item.get("college"),
                    "Priority Score": item.get("priority_score", 0.0),
                    "Reason": humanize_event_text(item.get("details", "")),
                }
            )

    return {
        "current_event": chunk[-1],
        "assignments": assignments,
        "waiting": waiting,
        "processed_count": len(chunk),
        "assigned_count": len(assignments),
        "waiting_count": len(waiting),
        "staff_flow": staff_flow,
    }


def on_playback_slider_change():
    """Sync slider position to internal request-step state and pause autoplay."""
    st.session_state.playback_playing = False
    st.session_state.playback_frame = max(0, int(st.session_state.playback_frame_ui) - 1)


def staff_rows_with_day_separators(rows: List[Dict]) -> List[Dict]:
    """Keep full staff history and visually separate each assignment day."""
    if not rows:
        return []

    ordered_rows = sorted(
        rows,
        key=lambda item: item.get("_dt") if item.get("_dt") is not None else parse_event_time(str(item.get("Assigned At", ""))),
    )

    display_rows: List[Dict] = []
    last_day = None
    day_block = 0
    day_count = 0

    for row in ordered_rows:
        assigned_day = row.get("_date")
        if assigned_day is None:
            assigned_at_raw = row.get("Assigned At")
            assigned_at_dt = parse_event_time(str(assigned_at_raw)) if assigned_at_raw else None
            assigned_day = assigned_at_dt.date() if assigned_at_dt else None

        if assigned_day != last_day:
            if last_day is not None:
                divider_text = f"--- Day {day_block + 1} Start ({format_compact_day(assigned_day)}) ---"
                display_rows.append(
                    {
                        "Day Block": divider_text,
                        "Count": "",
                        "Request": "",
                        "College": "",
                        "Document": "",
                        "Priority Score": "",
                        "Queue Wait (h)": "",
                        "Assigned At": "",
                    }
                )
            day_block += 1
            day_label = f"Day {day_block} ({format_compact_day(assigned_day)})"
            day_count = 0
        else:
            day_label = ""

        day_count += 1
        display_rows.append(
            {
                "Day Block": day_label,
                "Count": day_count,
                "Request": row.get("Request", ""),
                "College": row.get("College", ""),
                "Document": row.get("Document", ""),
                "Priority Score": row.get("Priority Score", ""),
                "Queue Wait (h)": row.get("Queue Wait (h)", ""),
                "Assigned At": format_compact_datetime(row.get("_dt") if row.get("_dt") is not None else row.get("Assigned At", "")),
            }
        )
        last_day = assigned_day

    return display_rows


initialize_state()


# ============================================================================
# SIDEBAR CONTROLS
# ============================================================================
@st.cache_data(ttl=10)
def fetch_custom_requests():
    """Cached HTTP fetch — re-fires at most every 10 s, not on every widget interaction."""
    try:
        res = requests.get(f"{BACKEND_URL}/api/custom-requests", timeout=5)
        if res.status_code == 200:
            return res.json()
    except Exception:
        pass
    return []


@st.fragment
def render_custom_request_manager():
    """Isolated fragment: only this section rerenders on add/delete/clear actions."""
    custom_reqs = fetch_custom_requests()

    with st.expander("(Postman Guide)", expanded=False):
        st.code(
            """
    POST http://localhost:5000/api/custom-requests

    Headers:
    Content-Type: application/json

    Body (raw JSON):
    {
    "college": "COE",
    "document_type": "Transcript of Records",
    "urgency": 9,
    "requester_type": "Graduating Student",
    "submission_time": "09:15",
    "payment_status": "Paid",
    "requirements_stage": "complete"
    }
    """,
            language="bash"
        )

    st.markdown("### ➕ Add Custom Request")
    c_college = st.selectbox("College", COLLEGES, key="c_req_college")
    c_doc = st.selectbox("Document Type", list(DOCUMENT_COMPLEXITY.keys()), key="c_req_doc")
    c_requester = st.selectbox("Requester Type", list(REQUESTER_PRIORITY.keys()), key="c_req_requester")
    c_urgency = st.slider("Urgency Level", min_value=1, max_value=10, value=5, key="c_req_urgency")
    c_sub_time = st.text_input("Submission Time", value="09:00", help="Use HH:MM format (e.g. 09:15) or full ISO datetime.", key="c_req_sub")
    c_payment = st.selectbox("Payment Status", ["Paid", "Unpaid"], key="c_req_payment")
    c_stage = st.selectbox("Requirements Stage", ["complete", "partial", "incomplete"], key="c_req_stage", format_func=lambda v: v.title())

    if st.button("➕ Add Request", use_container_width=True, key="c_req_add_btn"):
        payload = {
            "college": c_college,
            "document_type": c_doc,
            "urgency": c_urgency,
            "requester_type": c_requester,
            "submission_time": c_sub_time,
            "payment_status": c_payment,
            "requirements_stage": c_stage,
            "completeness_of_requirements": 1.0 if c_stage == "complete" else (0.7 if c_stage == "partial" else 0.3)
        }
        try:
            add_res = requests.post(f"{BACKEND_URL}/api/custom-requests", json=payload, timeout=5)
            if add_res.status_code == 201:
                st.success(f"Successfully added custom request: {add_res.json().get('request_id')}")
                fetch_custom_requests.clear()
                st.rerun(scope="fragment")
            else:
                st.error(f"Failed to add: {add_res.text}")
        except Exception as e:
            st.error(f"Error connecting to backend: {e}")

    if custom_reqs:
        st.markdown("### 📋 Existing Custom Requests")
        custom_df = pd.DataFrame([
            {
                "ID": r["request_id"],
                "College": r["college"],
                "Document": r["document_type"],
                "Urgency": r["urgency"],
                "Requester": r["requester_type"],
                "Submission": r["submission_time"],
                "Requirements": r["requirements_stage"].title(),
                "Payment": r["payment_status"]
            } for r in custom_reqs
        ])
        st.dataframe(custom_df, use_container_width=True)

        to_delete = st.selectbox("Select Request ID to Delete", options=[r["request_id"] for r in custom_reqs], key="c_req_to_delete")
        if st.button("🗑️ Delete Selected Request", use_container_width=True, key="c_req_del_btn"):
            try:
                del_res = requests.delete(f"{BACKEND_URL}/api/custom-requests/{to_delete}", timeout=5)
                if del_res.status_code == 200:
                    st.success(f"Deleted {to_delete}")
                    fetch_custom_requests.clear()
                    st.rerun(scope="fragment")
                else:
                    st.error(f"Failed to delete: {del_res.text}")
            except Exception as e:
                st.error(f"Error: {e}")

        if st.button("💥 Clear All Requests", use_container_width=True, key="c_req_clear_btn"):
            try:
                clear_res = requests.delete(f"{BACKEND_URL}/api/custom-requests", timeout=5)
                if clear_res.status_code == 200:
                    st.success("Cleared all custom requests")
                    fetch_custom_requests.clear()
                    st.rerun(scope="fragment")
                else:
                    st.error(f"Failed to clear: {clear_res.text}")
            except Exception as e:
                st.error(f"Error: {e}")
    else:
        st.info("No custom requests in the database. Add one above or via API. http://localhost:5000/api/custom-requests")


def _source_options(source_field: str) -> List[Any]:
    options = CRITERIA_SOURCE_OPTIONS.get(source_field)
    if options:
        return options
    if source_field == "requester_type":
        return list(REQUESTER_PRIORITY.keys())
    if source_field == "document_type":
        return list(DOCUMENT_COMPLEXITY.keys())
    if source_field == "college":
        return COLLEGES
    if source_field == "urgency":
        return list(range(1, 11))
    return []


def render_custom_request_fields_manager():
    st.caption(
        "Define generated request fields here. Option scores are normalized influence values from 0.00 to 1.00."
    )

    if CUSTOM_REQUEST_FIELDS:
        field_df = pd.DataFrame([
            {
                "Key": field.get("key"),
                "Name": field.get("label"),
                "Options": ", ".join(
                    f"{option.get('label')}={float(option.get('score', 0.0)):.2f}"
                    for option in field.get("options", [])
                ),
            }
            for field in CUSTOM_REQUEST_FIELDS
        ])
        st.dataframe(field_df, use_container_width=True, hide_index=True)
        delete_field_key = st.selectbox(
            "Delete request field",
            [field.get("key") for field in CUSTOM_REQUEST_FIELDS],
            format_func=lambda key: CRITERIA_SOURCE_FIELDS.get(key, key),
            key="delete_request_field_key",
        )
        if st.button("Delete Field", key="delete_request_field_btn", use_container_width=True):
            try:
                response = requests.delete(f"{BACKEND_URL}/api/request-fields/{delete_field_key}", timeout=5)
                if response.status_code == 200:
                    fetch_backend_config.clear()
                    reset_composition_group(delete_field_key)
                    st.success(f"Deleted {CRITERIA_SOURCE_FIELDS.get(delete_field_key, delete_field_key)}")
                    st.rerun()
                else:
                    st.error(response.text)
            except Exception as exc:
                st.error(f"Error deleting request field: {exc}")
    else:
        st.info("No custom request fields yet.")

    with st.form("custom_request_field_form"):
        field_label = st.text_input("Field Name", key="request_field_label_input")
        field_key = st.text_input(
            "Field Key",
            key="request_field_key_input",
            help="Optional. Leave blank to generate from the field name, e.g. Request Reason -> request_reason.",
        )
        st.caption("Category options and their normalized influence scores.")
        option_rows = st.number_input(
            "Number of Options",
            min_value=1,
            max_value=12,
            value=3,
            step=1,
            key="request_field_option_count",
        )
        options = []
        for index in range(int(option_rows)):
            label_col, score_col = st.columns([0.68, 0.32])
            option_label = label_col.text_input(
                f"Option {index + 1}",
                key=f"request_field_option_label_{index}",
            )
            option_score = score_col.number_input(
                "Score",
                min_value=0.0,
                max_value=1.0,
                value=1.0 if index == 0 else 0.0,
                step=0.05,
                format="%.2f",
                key=f"request_field_option_score_{index}",
            )
            if option_label.strip():
                options.append({"label": option_label.strip(), "score": float(option_score)})

        submitted = st.form_submit_button("Add Field", use_container_width=True)

    if submitted:
        payload = {
            "label": field_label,
            "key": field_key,
            "type": "category",
            "options": options,
        }
        try:
            response = requests.post(f"{BACKEND_URL}/api/request-fields", json=payload, timeout=5)
            if response.status_code == 201:
                fetch_backend_config.clear()
                st.success("Request field added. It now has its own composition mix.")
                st.rerun()
            else:
                st.error(response.text)
        except Exception as exc:
            st.error(f"Error adding request field: {exc}")


def render_custom_criteria_manager():
    st.caption("Link a custom request field into the weighted priority model.")

    if CUSTOM_CRITERIA:
        criteria_df = pd.DataFrame([
            {
                "Key": item.get("key"),
                "Name": item.get("label"),
                "Field": CRITERIA_SOURCE_FIELDS.get(item.get("source_field"), item.get("source_field")),
                "Rule": "Request field score",
            }
            for item in CUSTOM_CRITERIA
        ])
        st.dataframe(criteria_df, use_container_width=True, hide_index=True)
        delete_key = st.selectbox(
            "Delete custom criterion",
            [item.get("key") for item in CUSTOM_CRITERIA],
            format_func=lambda key: format_criterion_label(key),
            key="delete_custom_criterion_key",
        )
        if st.button("Delete Criterion", key="delete_custom_criterion_btn", use_container_width=True):
            try:
                response = requests.delete(f"{BACKEND_URL}/api/criteria/{delete_key}", timeout=5)
                if response.status_code == 200:
                    fetch_backend_config.clear()
                    st.success(f"Deleted {format_criterion_label(delete_key)}")
                    st.rerun()
                else:
                    st.error(response.text)
            except Exception as exc:
                st.error(f"Error deleting criterion: {exc}")
    else:
        st.info("No custom criteria yet.")

    linked_fields = {
        item.get("source_field")
        for item in CUSTOM_CRITERIA
    }
    available_custom_fields = [
        field
        for field in CUSTOM_REQUEST_FIELDS
        if field.get("key") not in linked_fields
    ]
    if not available_custom_fields:
        st.info("Add a custom request field first, or delete an existing linked criterion to relink it.")
        return

    with st.form("custom_criteria_form"):
        label = st.text_input("Criterion Name", key="criterion_label_input")
        source_field = st.selectbox(
            "Source Field",
            [field.get("key") for field in available_custom_fields],
            format_func=lambda key: CRITERIA_SOURCE_FIELDS.get(key, key),
            key="criterion_source_field",
        )
        st.caption("The selected request field's saved 0.00 to 1.00 option scores will be used directly.")

        submitted = st.form_submit_button("Add Criterion", use_container_width=True)

    if submitted:
        payload = {
            "label": label,
            "source_field": source_field,
            "scoring_type": "field_score",
        }

        try:
            response = requests.post(f"{BACKEND_URL}/api/criteria", json=payload, timeout=5)
            if response.status_code == 201:
                fetch_backend_config.clear()
                st.success("Custom criterion added. It will now appear in the criteria selector.")
                st.rerun()
            else:
                st.error(response.text)
        except Exception as exc:
            st.error(f"Error adding criterion: {exc}")


def render_composition_group(title: str, group: str, *, expanded: bool = False, disabled: bool = False):
    defaults = default_composition_values(group)
    with st.expander(title, expanded=expanded):
        if disabled:
            st.info("Disabled in Dataset Mode. Uploaded datasets already contain fixed request values.")
        st.caption(
            "Use quick target share for exact percentages, or edit the raw values manually below. "
            "Applied targets stay locked while the remaining share is distributed across unlocked options. "
            "Reset clears locked targets and returns this group to the currently active scenario defaults."
        )

        target_col, share_col, apply_col, reset_col = st.columns([0.40, 0.24, 0.18, 0.18])
        target_option = target_col.selectbox(
            "Target option",
            list(defaults.keys()),
            key=f"{group}_target_option",
            disabled=disabled,
        )
        target_share = share_col.number_input(
            "Target share",
            min_value=0.0,
            max_value=100.0,
            value=float(st.session_state.get(f"{group}_target_share_pct", 80.0)),
            step=1.0,
            format="%.1f",
            key=f"{group}_target_share_pct",
            help="Example: 80 means this option receives 80% of generated requests and the remaining 20% is distributed to the others.",
            disabled=disabled,
        )
        if apply_col.button("Apply", key=f"apply_{group}_target", use_container_width=True, disabled=disabled):
            set_composition_share(group, target_option, target_share / 100.0)
            st.rerun(scope="fragment")
        if reset_col.button("Reset", key=f"reset_{group}_composition", use_container_width=True, disabled=disabled):
            reset_composition_group(group)
            st.rerun(scope="fragment")

        locked_targets = get_composition_locked_targets(group)
        if locked_targets:
            st.caption(
                "Locked targets: "
                + ", ".join(
                    f"{option}={share:.1%}"
                    for option, share in locked_targets.items()
                )
            )

        total = 0.0
        for option, default_value in defaults.items():
            state_key = composition_state_key(group, option)
            total += float(st.number_input(
                str(option),
                min_value=0.0,
                value=float(st.session_state.get(state_key, default_value)),
                step=0.01,
                format="%.4f",
                key=state_key,
                disabled=disabled,
            ))
        if total <= 0:
            st.warning("Set at least one value above 0. If all values are 0, the backend falls back to defaults.")
        else:
            normalized_values = {
                option: float(st.session_state.get(composition_state_key(group, option), 0.0)) / total
                for option in defaults.keys()
            }
            st.caption(
                "Generated mix: "
                + ", ".join(
                    f"{option}={share:.1%}"
                    for option, share in normalized_values.items()
                    if share > 0
                )
            )


@st.fragment
def render_sidebar_controls():
    st.header("🎛️ Simulation Controls")

    pending_preset = st.session_state.pop("_pending_preset_to_load", None)
    if isinstance(pending_preset, dict):
        apply_ui_config(pending_preset)
        loaded_name = st.session_state.pop("_pending_preset_name", "selected preset")
        st.success(f"Loaded preset: {loaded_name}")

    run_col, reset_col = st.columns(2)
    run_clicked = run_col.button("🚀 Run", use_container_width=True)
    reset_clicked = reset_col.button("🧹 Reset", use_container_width=True)

    if reset_clicked:
        clear_run_state()
        st.rerun()

    st.selectbox(
        "Scheduler",
        SCHEDULER_OPTIONS,
        key="scheduler_type",
        format_func=lambda value: SCHEDULER_LABELS.get(value, value),
    )
    st.selectbox(
        "Allocator",
        ALLOCATOR_OPTIONS,
        key="allocator_type",
        format_func=lambda value: ALLOCATOR_LABELS.get(value, value.replace("_", " ").title()),
    )
    if st.session_state.scheduler_type == "WEIGHTED":
        with st.expander("📊 Priority Weights", expanded=False):

            st.subheader("Weighted Priority")
            st.caption(
                "Gate checks: "
                + ", ".join(format_criterion_label(key) for key in GATE_CRITERIA_KEYS)
                + ". These must be ready before assignment and are not weighted."
            )
            with st.expander("Custom Criteria", expanded=False):
                render_custom_criteria_manager()

            available_criteria = active_criteria()
            sync_selected_weight_criteria()
            st.multiselect(
                "Criteria included in priority score",
                available_criteria,
                default=st.session_state.selected_weight_criteria,
                key="selected_weight_criteria",
                format_func=format_criterion_label,
            )

            st.radio(
                "Weighting mode",
                list(WEIGHT_MODE_OPTIONS.keys()),
                key="weight_mode",
                format_func=lambda value: WEIGHT_MODE_OPTIONS[value],
                horizontal=True,
            )

            selected_criteria = get_ordered_selected_criteria()
            if st.session_state.weight_mode == "roc":
                st.caption("Rank criteria from highest to lowest priority. ROC weights are calculated from this order.")
                ordered_criteria = get_criteria_order()
                for index, key in enumerate(ordered_criteria):
                    rank_col, label_col, up_col, down_col = st.columns([0.10, 0.70, 0.10, 0.10], vertical_alignment="center")
                    rank_col.markdown(f"**{index + 1}.**")
                    label_col.markdown(format_criterion_label(key))
                    if up_col.button("↑", key=f"move_up_{key}", disabled=index == 0, help="Move higher"):
                        ordered_criteria[index - 1], ordered_criteria[index] = ordered_criteria[index], ordered_criteria[index - 1]
                        set_criteria_order(ordered_criteria)
                        st.rerun(scope="fragment")
                    if down_col.button("↓", key=f"move_down_{key}", disabled=index == len(ordered_criteria) - 1, help="Move lower"):
                        ordered_criteria[index + 1], ordered_criteria[index] = ordered_criteria[index], ordered_criteria[index + 1]
                        set_criteria_order(ordered_criteria)
                        st.rerun(scope="fragment")
            else:
                reconcile_manual_weight_state()
                st.caption("Manual weights use 0.00 to 1.00 values and cannot exceed a total of 1.00.")
                for key in selected_criteria:
                    state_key = weight_state_key(key)
                    current_value = max(0.0, min(float(st.session_state.get(state_key, 0.0)), 1.0))
                    other_total = sum(
                        max(0.0, min(float(st.session_state.get(weight_state_key(other_key), 0.0)), 1.0))
                        for other_key in selected_criteria
                        if other_key != key
                    )
                    max_allowed = max(0.0, min(1.0, 1.0 - other_total))
                    disabled = current_value <= 0.0 and max_allowed <= 0.001
                    if current_value > max_allowed:
                        current_value = max_allowed
                    st.session_state[state_key] = current_value
                    if disabled:
                        st.session_state[state_key] = 0.0
                    st.slider(
                        f"Weight: {format_criterion_label(key)}",
                        min_value=0.0,
                        max_value=1.0 if disabled else max(float(max_allowed), 0.01),
                        step=0.01,
                        key=state_key,
                        disabled=disabled,
                    )
                manual_total = sum(
                    float(st.session_state.get(weight_state_key(key), 0.0))
                    for key in selected_criteria
                )
                st.progress(min(manual_total, 1.0))
                st.caption(f"Manual total: {manual_total:.2f} / 1.00")
                if manual_total < 0.999:
                    st.warning(f"Assign the remaining {1.0 - manual_total:.2f} before running.")

            current_weights = normalized_weights_from_ui()
            weight_caption = "Configured" if st.session_state.get("weight_mode", "roc") == "manual" else "Normalized"
            st.caption(
                f"{weight_caption}: "
                + ", ".join(f"{format_criterion_label(k)}={v:.2f}" for k, v in current_weights.items()),
                help="Tie-break rule: earlier submission_time wins when scores are equal."
            )
    st.subheader("Capacity and Policy")
    st.slider(
        "Number of Staff",
        min_value=len(COLLEGES),
        max_value=len(COLLEGES) * 2,
        step=1,
        key="num_staff",
    )

    max_absent_staff = max(0, int(st.session_state.num_staff) - 1)
    st.checkbox(
        "🤒 Enable Staff Absence",
        key="enable_absence",
        disabled=(max_absent_staff == 0),
        help="Turn on to model staff being absent during the run.",
    )

    if max_absent_staff == 0:
        st.session_state.enable_absence = False
        st.session_state.num_absent_staff = 0
    elif st.session_state.enable_absence:
        if st.session_state.num_absent_staff < 1:
            st.session_state.num_absent_staff = 1
        if st.session_state.num_absent_staff > max_absent_staff:
            st.session_state.num_absent_staff = max_absent_staff

        st.slider(
            "Number of Absent Staff",
            min_value=1,
            max_value=max_absent_staff,
            step=1,
            key="num_absent_staff",
        )
    else:
        st.session_state.num_absent_staff = 0

    st.slider("Daily Quota per Staff", min_value=1, max_value=60, step=1, key="quota_limit")
    st.time_input("Workday Start", key="work_start_time")
    st.time_input("Workday End", key="work_end_time")

    st.subheader("📥 Demand")

    if "total_requests" not in st.session_state:
        st.session_state.total_requests = DEFAULT_STATE["total_requests"]
    if "disable_generated_requests" not in st.session_state:
        st.session_state.disable_generated_requests = DEFAULT_STATE["disable_generated_requests"]

    def on_peak_mode_change():
        if st.session_state.get("peak_mode", False):
            st.session_state.previous_total_requests = st.session_state.get("total_requests", DEFAULT_STATE["total_requests"])
            st.session_state.total_requests = 300
        else:
            if "previous_total_requests" in st.session_state:
                st.session_state.total_requests = (
                    st.session_state.previous_total_requests
                )
        reset_composition_group("document_type")

    with st.expander("Request Dataset CSV", expanded=False):
        uploaded_dataset_file = st.file_uploader(
            "Upload Request Dataset CSV",
            type=["csv"],
            key="request_dataset_uploader",
            help="Uses this CSV for the run only. It is not saved into the custom request database.",
        )
        if uploaded_dataset_file is not None:
            try:
                rows = parse_request_dataset_csv(uploaded_dataset_file)
                st.session_state.uploaded_request_dataset = rows
                st.session_state.uploaded_request_dataset_name = uploaded_dataset_file.name
                st.session_state.disable_generated_requests = True
                st.session_state.peak_mode = False
                st.success(f"Loaded {len(rows)} requests from {uploaded_dataset_file.name}")
            except Exception as exc:
                st.session_state.uploaded_request_dataset = None
                st.session_state.uploaded_request_dataset_name = None
                st.error(f"Could not load dataset: {exc}")

        if st.session_state.get("uploaded_request_dataset"):
            st.caption(
                f"Active dataset: {st.session_state.get('uploaded_request_dataset_name')} "
                f"({len(st.session_state.uploaded_request_dataset)} requests). Generated requests will be disabled for the next run."
            )
            if st.button("Clear Uploaded Dataset", key="clear_uploaded_request_dataset", use_container_width=True):
                st.session_state.uploaded_request_dataset = None
                st.session_state.uploaded_request_dataset_name = None
                st.session_state.disable_generated_requests = DEFAULT_STATE["disable_generated_requests"]
                st.rerun(scope="fragment")

    dataset_mode = bool(st.session_state.get("uploaded_request_dataset"))
    if dataset_mode:
        st.session_state.disable_generated_requests = True
        st.session_state.peak_mode = False
        st.info(
            "Dataset Mode is active. Request generation controls are disabled because "
            "the uploaded CSV is the source of requests for the next run. Clear the "
            "uploaded dataset to re-enable generated-request settings."
        )

    st.checkbox(
        "Disable Generated Requests",
        key="disable_generated_requests",
        disabled=dataset_mode,
        help=(
            "Locked on while Dataset Mode is active. Clear the uploaded dataset to enable generated requests again."
            if dataset_mode
            else "If enabled, only custom requests stored in the database are simulated."
        ),
    )

    st.slider(
        "Total Daily Requests",
        min_value=50,
        max_value=500,
        step=10,
        key="total_requests",
        disabled=(
            st.session_state.get("peak_mode", False)
            or st.session_state.get("disable_generated_requests", False)
            or dataset_mode
        ),
    )

    st.subheader("Custom Requests")
    with st.expander("🛠️ Custom Request Manager", expanded=False):
        render_custom_request_manager()
    st.subheader("Features")
    st.checkbox("⚡ Enable Urgency", key="urgency")

    st.checkbox(
        "📈 Peak Period",
        key="peak_mode",
        on_change=on_peak_mode_change,
        disabled=dataset_mode,
        help="Disabled in Dataset Mode because the uploaded dataset already defines request demand." if dataset_mode else None,
    )
    st.subheader("Request Composition")
    st.caption(
        "Composition is the source of truth for generated request likelihood. "
        "College Mix controls college demand directly. Normal documents are equal; "
        "Peak Period boosts TOR/TC and Certification."
    )
    with st.expander("Custom Request Fields", expanded=False):
        render_custom_request_fields_manager()
    render_composition_group("College Mix", "college", expanded=False, disabled=dataset_mode)
    render_composition_group("Requester Type Mix", "requester_type", expanded=False, disabled=dataset_mode)
    render_composition_group("Document Type Mix", "document_type", expanded=False, disabled=dataset_mode)
    for field in CUSTOM_REQUEST_FIELDS:
        field_key = field.get("key")
        if field_key:
            render_composition_group(
                f"{field.get('label', field_key)} Mix",
                str(field_key),
                expanded=False,
                disabled=dataset_mode,
            )

    st.subheader("Presets")
    presets = load_presets()
    preset_names = ["(select)"] + sorted(list(presets.keys()))
    selected_preset = st.selectbox("Saved Presets", preset_names)

    load_col, save_col, remove_col = st.columns(3)
    load_clicked = load_col.button("Load", use_container_width=True)
    save_clicked = save_col.button("Save", use_container_width=True)
    remove_clicked = remove_col.button(
        "Remove",
        use_container_width=True,
        disabled=selected_preset not in presets,
    )

    preset_name_input = st.text_input("Preset Name", value="")

    if load_clicked and selected_preset in presets:
        st.session_state["_pending_preset_to_load"] = presets[selected_preset]
        st.session_state["_pending_preset_name"] = selected_preset
        st.rerun()

    if remove_clicked and selected_preset in presets:
        removed_name = selected_preset
        presets.pop(removed_name, None)
        save_presets(presets)
        st.success(f"Removed preset: {removed_name}")
        st.rerun(scope="fragment")

    if save_clicked:
        name = preset_name_input.strip()
        if name:
            presets[name] = collect_ui_config()
            save_presets(presets)
            st.success(f"Saved preset: {name}")
        else:
            st.warning("Enter a preset name before saving.")

    if st.session_state.simulation_engine is not None:
        with st.expander("🐛 Debug: Urgency Status", expanded=False):
            st.markdown(f"**Checkbox State:** `{st.session_state.urgency}`")
            st.markdown(f"**Current Weight for Urgency:** `{normalized_weights_from_ui().get('urgency', 'N/A')}`")
                
            if st.session_state.simulation_results and st.session_state.simulation_results.get('completed_requests'):
                sample = st.session_state.simulation_results['completed_requests'][0]
                st.markdown(f"**Sample Request `{sample['request_id']}` Priority:** `{sample['priority_score']}`")
                st.caption("Run twice (checkbox OFF/ON) to compare this number.")

    if run_clicked:
        run_simulation_now()
        if st.session_state.simulation_results is not None:
            st.rerun()


with st.sidebar:
    render_sidebar_controls()

# ============================================================================
# MAIN RESULTS
# ============================================================================

if not st.session_state.simulation_results or not st.session_state.simulation_engine:
    st.info("Use the sidebar controls, then click Run to start the simulation.")
    st.stop()

engine = st.session_state.simulation_engine
results = st.session_state.simulation_results
is_weighted_scheduler = results.get("scheduler_type") == "WEIGHTED"
staff_college_map = build_staff_college_map(engine.staff_pool)

st.success("Simulation complete.")

run_cfg = results.get("run_config", {})

num_staff = run_cfg.get("num_staff")
if num_staff is None:
    num_staff = len(results.get("staff_info", [])) or st.session_state.num_staff

quota_limit = run_cfg.get("quota_limit")
if quota_limit is None:
    staff_info = results.get("staff_info", [])
    quota_limit = staff_info[0].get("quota_limit") if staff_info else st.session_state.quota_limit

num_absent = run_cfg.get("num_absent_staff", 0)
if num_absent > 0:
    absence_status = f"🔴 Enabled ({num_absent} absent)"
else:
    absence_status = "🟢 None (All active)"

disable_generated = run_cfg.get("disable_generated_requests", False)
all_reqs = results.get("generated_requests", [])
total_custom = sum(1 for r in all_reqs if r.get("is_custom"))
total_gen = sum(1 for r in all_reqs if not r.get("is_custom"))

if run_cfg.get("uploaded_request_dataset_mode"):
    demand_mode = f"Uploaded Dataset ({len(all_reqs)} reqs)"
elif disable_generated:
    demand_mode = f"⭐ Custom only ({total_custom} reqs)"
else:
    demand_mode = f"🤖 Gen: {total_gen} | ⭐ Cust: {total_custom}"

urgency_enabled = run_cfg.get("urgency", False)
urgency_status = "⚡ Enabled" if urgency_enabled else "❌ Disabled"

scenario = run_cfg.get("scenario", "baseline")
is_peak = (scenario == "peak_period" or run_cfg.get("peak_mode", False))
peak_status = "🔥 Peak" if is_peak else "Regular"


scheduler_label = SCHEDULER_LABELS.get(results.get('scheduler_type'), results.get('scheduler_type'))
allocator_label = ALLOCATOR_LABELS.get(results.get('allocator_type'), str(results.get('allocator_type')).replace('_', ' ').title())

st.markdown(
    f"""
    <div class="hero-band" style="flex-direction: column; align-items: stretch; gap: 0.8rem;">
        <div style="display: flex; justify-content: space-between; align-items: center; width: 100%;">
            <div>
                <p class="hero-kicker">Simulation Snapshot</p>
                <p class="hero-title">Scheduler: {scheduler_label} | Allocator: {allocator_label}</p>
                <p class="hero-sub">Mode: Configurable request composition</p>
            </div>
            <div class="hero-pill">Ready for playback</div>
        </div>
        <div class="snap-grid">
            <div class="snap-item">
                <span class="snap-label">👥 Staffing</span>
                <span class="snap-value">{num_staff - num_absent} Staff (Quota: {quota_limit}/day)</span>
            </div>
            <div class="snap-item">
                <span class="snap-label">🤒 Staff Absence</span>
                <span class="snap-value">{absence_status}</span>
            </div>
            <div class="snap-item">
                <span class="snap-label">📥 Demand Mode</span>
                <span class="snap-value">{demand_mode}</span>
            </div>
            <div class="snap-item">
                <span class="snap-label">⚡ Urgency State</span>
                <span class="snap-value">{urgency_status}</span>
            </div>
            <div class="snap-item">
                <span class="snap-label">📈 Peak Period</span>
                <span class="snap-value">{peak_status}</span>
            </div>        </div>
    </div>
    """,
    unsafe_allow_html=True,
)


@st.fragment
def render_playback_section(results, engine):
    st.header("Playback")

    event_log = results.get("event_log", [])
    
    run_key = (results.get("seed_used"), results.get("scheduler_type"), results.get("allocator_type"))
    if st.session_state.get("cached_decisions_run_key") != run_key:
        st.session_state.decisions = routing_events(event_log)
        st.session_state.cached_decisions_run_key = run_key
    decisions = st.session_state.decisions

    if not decisions:
        st.warning("No request-routing decisions available for playback.")
        return

    max_step = len(decisions) - 1
    st.session_state.playback_frame = min(st.session_state.playback_frame, max_step)
    st.session_state.playback_frame_ui = min(max(st.session_state.playback_frame_ui, 1), max_step + 1)

    force_slider_sync = False

    controls_col1, controls_col2, controls_col3, controls_col4, controls_col5, controls_col6 = st.columns(6)

    if controls_col1.button("Play", use_container_width=True):
        st.session_state.playback_playing = True
    if controls_col2.button("Pause", use_container_width=True):
        st.session_state.playback_playing = False
    if controls_col3.button("Step Back", use_container_width=True):
        st.session_state.playback_playing = False
        st.session_state.playback_frame = max(0, st.session_state.playback_frame - 1)
        force_slider_sync = True
    if controls_col4.button("Step Forward", use_container_width=True):
        st.session_state.playback_playing = False
        st.session_state.playback_frame = min(max_step, st.session_state.playback_frame + 1)
        force_slider_sync = True
    if controls_col5.button("End", use_container_width=True):
        st.session_state.playback_playing = False
        st.session_state.playback_frame = max_step
        force_slider_sync = True

    controls_col6.selectbox("Speed", list(SPEED_OPTIONS.keys()), key="playback_speed")

    if force_slider_sync or (
        st.session_state.playback_playing
        and st.session_state.playback_frame_ui != st.session_state.playback_frame
    ):
        st.session_state.playback_frame_ui = st.session_state.playback_frame + 1

    st.slider(
        "Request Step",
        min_value=1,
        max_value=max_step + 1,
        key="playback_frame_ui",
        on_change=on_playback_slider_change,
    )

    frame_data = playback_state(decisions, st.session_state.playback_frame)
    current_event = frame_data["current_event"]

    if st.session_state.get("cached_request_lookup_run_key") != run_key:
        lookup = {}
        for request_item in results.get("generated_requests", []):
            if isinstance(request_item, dict) and request_item.get("request_id"):
                req_copy = request_item.copy()
                sub_raw = req_copy.get("submission_time")
                ready_raw = req_copy.get("ready_time")
                req_copy["_submission_time_parsed"] = parse_event_time(sub_raw) if sub_raw else None
                req_copy["_ready_time_parsed"] = parse_event_time(ready_raw) if ready_raw else req_copy["_submission_time_parsed"]
                lookup[req_copy["request_id"]] = req_copy
        st.session_state.request_lookup = lookup
        st.session_state.cached_request_lookup_run_key = run_key
    request_lookup = st.session_state.request_lookup

    staff_rows: Dict[str, List[Dict]] = {}
    staff_meta: Dict[str, Dict] = {}
    for staff in engine.staff_pool:
        staff_rows[staff.staff_id] = []
        staff_meta[staff.staff_id] = {
            "college": staff.college_affiliation,
            "quota": staff.quota_limit,
        }

    for assignment in frame_data["assignments"]:
        staff_id = assignment.get("Staff") or "UNASSIGNED"
        request_id = assignment.get("Request")
        request_meta = request_lookup.get(request_id, {})
        priority_score = request_meta.get("priority_score", assignment.get("Priority Score", 0.0))

        if staff_id not in staff_rows:
            staff_rows[staff_id] = []
            staff_meta[staff_id] = {"college": "-", "quota": None}

        is_custom_req = request_meta.get("is_custom", False)
        req_display = f"⭐ {request_id}" if is_custom_req else request_id

        assigned_at_raw = assignment.get("Time")
        assigned_at_dt = parse_event_time(str(assigned_at_raw)) if assigned_at_raw else None
        assigned_at_date = assigned_at_dt.date() if assigned_at_dt else None

        staff_rows[staff_id].append(
            {
                "Request": req_display,
                "College": request_meta.get("college", assignment.get("College")),
                "Document": request_meta.get("document_type", "-"),
                "Priority Score": round(float(priority_score or 0.0), 4),
                "Queue Wait (h)": assignment.get("Queue Wait (h)"),
                "Assigned At": assignment.get("Time"),
                "_dt": assigned_at_dt,
                "_date": assigned_at_date,
            }
        )

    assigned_request_ids = {
        assignment.get("Request")
        for assignment in frame_data["assignments"]
        if assignment.get("Request")
    }

    is_weighted_scheduler = results.get("scheduler_type") == "WEIGHTED"
    staff_college_map = build_staff_college_map(engine.staff_pool)

    if current_event:
        current_time = parse_event_time(current_event.get("time", ""))

        routed_request_ids = set()
        for assign_item in frame_data["assignments"]:
            if assign_item.get("Request"):
                routed_request_ids.add(assign_item["Request"])
        req_display = f"⭐ {request_id}" if is_custom_req else request_id
        
        waiting_rows.append(
            {
                "Request": req_display,
                "College": request_meta.get("college", waiting_item.get("College")),
                "Document": request_meta.get("document_type", "-"),
                "Priority Score": round(float(request_meta.get("priority_score", waiting_item.get("Priority Score", 0.0)) or 0.0), 4),
                "Submitted": format_compact_datetime(request_meta.get("submission_time", "-")),
                "Reason": waiting_item.get("Reason", ""),
                "Event Time": format_compact_datetime(event_time_raw),
                "_event_time": parse_event_time(str(event_time_raw)) if event_time_raw else datetime.min,
            }
        )

    is_weighted_scheduler = results.get("scheduler_type") == "WEIGHTED"
    staff_college_map = build_staff_college_map(engine.staff_pool)

    if is_weighted_scheduler:
        waiting_rows.sort(
            key=lambda row: (
                -float(row.get("Priority Score", 0.0)),
                row.get("_event_time", datetime.min),
            )
        )

    if current_event:
        current_time = parse_event_time(current_event.get("time", ""))

        routed_request_ids = set()
        for assign_item in frame_data["assignments"]:
            if assign_item.get("Request"):
                routed_request_ids.add(assign_item["Request"])

        pending_queue_rows = []
        for request_id, request_meta in request_lookup.items():
            submission_time = request_meta.get("_submission_time_parsed")
            if not submission_time:
                continue
            if submission_time <= current_time and request_id not in routed_request_ids:
                is_custom_req = request_meta.get("is_custom", False)
                req_display = f"⭐ {request_id}" if is_custom_req else request_id
                pending_queue_rows.append(
                    {
                        "Request": req_display,
                        "College": request_meta.get("college", "-"),
                        "Document": request_meta.get("document_type", "-"),
                        "Priority Score": round(float(request_meta.get("priority_score", 0.0) or 0.0), 4),
                        "Submitted": format_compact_datetime(request_meta.get("submission_time")),
                        "Pending Wait (h)": round((current_time - submission_time).total_seconds() / 3600.0, 2),
                        "_sort_submission": submission_time,
                    }
                )

        if is_weighted_scheduler:
            pending_queue_rows.sort(
                key=lambda row: (-float(row.get("Priority Score", 0.0)), row["_sort_submission"])
            )
        else:
            pending_queue_rows.sort(key=lambda row: row["_sort_submission"])
        for row in pending_queue_rows:
            row.pop("_sort_submission", None)

        card1, card2, card3, card4, card5 = st.columns([1.6, 1, 1, 1, 1])
        card1.metric("Simulation Clock", current_time.strftime("%Y-%m-%d %H:%M"))
        card2.metric("Current Request Step", f"{st.session_state.playback_frame + 1}/{max_step + 1}")
        card3.metric("Processed Decisions", frame_data["processed_count"])
        card4.metric("Assigned So Far", frame_data["assigned_count"])
        card5.metric("Queue Size Now", len(pending_queue_rows))

        routing_event_label = str(current_event.get("event_type", "")).replace("_", " ").title()
        routing_detail_label = str(current_event.get("details", "")).replace("_", " ").title()

        st.markdown(
            "**Current Routing Decision:** "
            f"{routing_event_label} | "
            f"Request={current_event.get('request_id')} | "
            f"Staff={format_staff_label(current_event.get('staff_id'), staff_college_map)} | "
            f"Details={routing_detail_label}"
        )

        st.subheader("Staff Capacity View")
        quota_enforced = results.get("allocator_type") != "quota_free"
        current_day = current_time.date()
        capacity_rows = []
        for staff in engine.staff_pool:
            rows_for_staff = staff_rows.get(staff.staff_id, [])
            total_assigned = len(rows_for_staff)
            assigned_today = 0
            for row in rows_for_staff:
                if row.get("_date") == current_day:
                    assigned_today += 1

            quota_value = staff.quota_limit if quota_enforced else None
            row = {
                "Staff ID": staff.staff_id,
                "Staff": format_staff_label(staff.staff_id, staff_college_map),
                "College": staff.college_affiliation,
                "Assigned Today": assigned_today,
                "Total Assigned": total_assigned,
            }
            if quota_enforced:
                row["Quota/Day"] = quota_value
                row["Today Fill %"] = round((assigned_today / max(quota_value, 1)) * 100.0, 1)
            capacity_rows.append(row)

        capacity_df = pd.DataFrame(capacity_rows)
        assigned_today_map = {
            row["Staff ID"]: row["Assigned Today"] for row in capacity_rows
        }
        total_assigned_map = {
            row["Staff ID"]: row["Total Assigned"] for row in capacity_rows
        }

        fig_capacity = go.Figure()
        fig_capacity.add_trace(
            go.Bar(
                name="Assigned Today",
                x=capacity_df["Staff"],
                y=capacity_df["Assigned Today"],
                marker_color="#a855f7",
                text=capacity_df["Assigned Today"],
                textposition="outside",
            )
        )
        if quota_enforced:
            fig_capacity.add_trace(
                go.Bar(
                    name="Quota per Day",
                    x=capacity_df["Staff"],
                    y=capacity_df["Quota/Day"],
                    marker_color="#22d3ee",
                    opacity=0.55,
                    text=capacity_df["Quota/Day"],
                    textposition="outside",
                )
            )

        fig_capacity.update_layout(
            title="Daily Staff Capacity at Current Request Step",
            xaxis_title="Staff",
            yaxis_title="Requests",
            barmode="group",
            height=360,
        )
        apply_plot_theme(fig_capacity)
        st.plotly_chart(fig_capacity, use_container_width=True, config={"staticPlot": True})

        st.subheader("Live Staff Request Lists")
        ordered_staff_ids = [staff.staff_id for staff in engine.staff_pool]
        for idx in range(0, len(ordered_staff_ids), 3):
            row_ids = ordered_staff_ids[idx: idx + 3]
            row_cols = st.columns(3)
            for col, staff_id in zip(row_cols, row_ids):
                with col:
                    meta = staff_meta.get(staff_id, {"college": "-", "quota": None})
                    assigned_today = assigned_today_map.get(staff_id, 0)
                    total_assigned = total_assigned_map.get(staff_id, 0)
                    if quota_enforced and meta.get("quota") is not None:
                        quota_value = int(meta["quota"])
                        st.metric(
                            f"{staff_id} ({meta['college']})",
                            f"{assigned_today}/{quota_value}",
                            delta=f"Total {total_assigned}",
                        )
                        if assigned_today >= quota_value:
                            st.warning("Quota full at this step")
                    else:
                        st.metric(
                            f"{staff_id} ({meta['college']})",
                            f"{total_assigned} assigned",
                        )

                    rows = staff_rows.get(staff_id, [])
                    if rows:
                        display_rows = staff_rows_with_day_separators(rows)
                        render_theme_table(pd.DataFrame(display_rows), height_px=340)
                    else:
                        st.caption("No requests routed here yet.")

        st.subheader("Pending Queue")
        st.caption("Requests already arrived but still waiting for a slot.")

        if pending_queue_rows:
            render_theme_table(pd.DataFrame(pending_queue_rows), height_px=320)
        else:
            st.caption("No pending queue at this step.")
        st.subheader("Unassignable Waiting List")
        st.caption("Requests that cannot be routed to any staff.")

        if waiting_rows:
            waiting_display_rows = [
                {k: v for k, v in row.items() if not str(k).startswith("_")}
                for row in waiting_rows
            ]
            render_theme_table(pd.DataFrame(waiting_display_rows), height_px=320)
        else:
            st.caption("No unassignable waiting requests at this step.")

    if st.session_state.playback_playing:
        if st.session_state.playback_frame < max_step:
            tm.sleep(SPEED_OPTIONS.get(st.session_state.playback_speed, 0.45))
            st.session_state.playback_frame += 1
            st.rerun(scope="fragment")
        else:
            st.session_state.playback_playing = False


@st.fragment
def render_playback_section(results, engine):
    st.header("Playback")

    event_log = results.get("event_log", [])
    run_key = (results.get("seed_used"), results.get("scheduler_type"), results.get("allocator_type"))
    if st.session_state.get("cached_decisions_run_key") != run_key:
        st.session_state.decisions = routing_events(event_log)
        st.session_state.cached_decisions_run_key = run_key
    decisions = st.session_state.decisions

    if not decisions:
        st.warning("No request-routing decisions available for playback.")
        return

    max_step = len(decisions) - 1
    st.session_state.playback_frame = min(st.session_state.playback_frame, max_step)
    st.session_state.playback_frame_ui = min(max(st.session_state.playback_frame_ui, 1), max_step + 1)

    force_slider_sync = False
    controls_col1, controls_col2, controls_col3, controls_col4, controls_col5, controls_col6 = st.columns(6)
    if controls_col1.button("Play", use_container_width=True, key="play_fixed"):
        st.session_state.playback_playing = True
    if controls_col2.button("Pause", use_container_width=True, key="pause_fixed"):
        st.session_state.playback_playing = False
    if controls_col3.button("Step Back", use_container_width=True, key="back_fixed"):
        st.session_state.playback_playing = False
        st.session_state.playback_frame = max(0, st.session_state.playback_frame - 1)
        force_slider_sync = True
    if controls_col4.button("Step Forward", use_container_width=True, key="step_fixed"):
        st.session_state.playback_playing = False
        st.session_state.playback_frame = min(max_step, st.session_state.playback_frame + 1)
        force_slider_sync = True
    if controls_col5.button("End", use_container_width=True, key="end_fixed"):
        st.session_state.playback_playing = False
        st.session_state.playback_frame = max_step
        force_slider_sync = True
    controls_col6.selectbox("Speed", list(SPEED_OPTIONS.keys()), key="playback_speed")

    if force_slider_sync or (
        st.session_state.playback_playing
        and st.session_state.playback_frame_ui != st.session_state.playback_frame
    ):
        st.session_state.playback_frame_ui = st.session_state.playback_frame + 1

    st.slider(
        "Request Step",
        min_value=1,
        max_value=max_step + 1,
        key="playback_frame_ui",
        on_change=on_playback_slider_change,
    )

    frame_data = playback_state(decisions, st.session_state.playback_frame)
    current_event = frame_data["current_event"]
    if not current_event:
        return

    if st.session_state.get("cached_request_lookup_run_key") != run_key:
        lookup = {}
        for request_item in results.get("generated_requests", []):
            if isinstance(request_item, dict) and request_item.get("request_id"):
                req_copy = request_item.copy()
                sub_raw = req_copy.get("submission_time")
                ready_raw = req_copy.get("ready_time")
                req_copy["_submission_time_parsed"] = parse_event_time(sub_raw) if sub_raw else None
                req_copy["_ready_time_parsed"] = parse_event_time(ready_raw) if ready_raw else req_copy["_submission_time_parsed"]
                lookup[req_copy["request_id"]] = req_copy
        st.session_state.request_lookup = lookup
        st.session_state.cached_request_lookup_run_key = run_key
    request_lookup = st.session_state.request_lookup

    staff_rows: Dict[str, List[Dict]] = {}
    staff_meta: Dict[str, Dict] = {}
    for staff in engine.staff_pool:
        staff_rows[staff.staff_id] = []
        staff_meta[staff.staff_id] = {"college": staff.college_affiliation, "quota": staff.quota_limit}

    for assignment in frame_data["assignments"]:
        staff_id = assignment.get("Staff") or "UNASSIGNED"
        request_id = assignment.get("Request")
        request_meta = request_lookup.get(request_id, {})
        if staff_id not in staff_rows:
            staff_rows[staff_id] = []
            staff_meta[staff_id] = {"college": "-", "quota": None}
        assigned_at_raw = assignment.get("Time")
        assigned_at_dt = parse_event_time(str(assigned_at_raw)) if assigned_at_raw else None
        staff_rows[staff_id].append(
            {
                "Request": request_id,
                "College": request_meta.get("college", assignment.get("College")),
                "Document": request_meta.get("document_type", "-"),
                "Priority Score": round(float(request_meta.get("priority_score", assignment.get("Priority Score", 0.0)) or 0.0), 4),
                "Queue Wait (h)": assignment.get("Queue Wait (h)"),
                "Assigned At": assignment.get("Time"),
                "_dt": assigned_at_dt,
                "_date": assigned_at_dt.date() if assigned_at_dt else None,
            }
        )

    current_time = parse_event_time(current_event.get("time", ""))
    routed_request_ids = {
        item.get("Request")
        for item in frame_data["assignments"]
        if item.get("Request")
    }
    is_weighted_scheduler = results.get("scheduler_type") == "WEIGHTED"
    staff_college_map = build_staff_college_map(engine.staff_pool)

    latest_rank_order: Dict[str, int] = {}
    latest_rank_scores: Dict[str, float] = {}
    if is_weighted_scheduler:
        priority_events = [
            event for event in event_log
            if event.get("event_type") == "PRIORITY"
            and parse_event_time(str(event.get("time", ""))) <= current_time
        ]
        if priority_events:
            latest_priority = priority_events[-1]
            latest_rank_order = {
                str(request_id): index
                for index, request_id in enumerate(latest_priority.get("ranked_request_ids") or [])
            }
            latest_rank_scores = {
                str(request_id): float(score or 0.0)
                for request_id, score in (latest_priority.get("ranked_scores") or {}).items()
            }

    pending_queue_rows = []
    not_ready_waiting_rows = []
    for request_id, request_meta in request_lookup.items():
        submission_time = request_meta.get("_submission_time_parsed")
        if not submission_time or submission_time > current_time or request_id in routed_request_ids:
            continue
        ready_time = request_meta.get("_ready_time_parsed") or submission_time
        base_row = {
            "Request": request_id,
            "College": request_meta.get("college", "-"),
            "Document": request_meta.get("document_type", "-"),
            "Priority Score": round(float(latest_rank_scores.get(request_id, request_meta.get("priority_score", 0.0)) or 0.0), 4),
            "Submitted": format_compact_datetime(request_meta.get("submission_time")),
            "_sort_submission": submission_time,
            "_rank": latest_rank_order.get(request_id, 999999),
        }
        if ready_time and current_time < ready_time:
            not_ready_waiting_rows.append(
                {
                    **base_row,
                    "Ready At": format_compact_datetime(request_meta.get("ready_time")),
                    "Waiting Reason": "Pending requirements or payment",
                    "Wait Until Ready (h)": round((ready_time - current_time).total_seconds() / 3600.0, 2),
                }
            )
        else:
            pending_queue_rows.append(
                {
                    **base_row,
                    "Ready Since": format_compact_datetime(request_meta.get("ready_time") or request_meta.get("submission_time")),
                    "Pending Wait (h)": round((current_time - ready_time).total_seconds() / 3600.0, 2),
                }
            )

    if is_weighted_scheduler:
        pending_queue_rows.sort(key=lambda row: (row.get("_rank", 999999), row["_sort_submission"]))
    else:
        pending_queue_rows.sort(key=lambda row: row["_sort_submission"])
    not_ready_waiting_rows.sort(key=lambda row: row["_sort_submission"])
    for row in pending_queue_rows + not_ready_waiting_rows:
        row.pop("_sort_submission", None)
        row.pop("_rank", None)

    card1, card2, card3, card4, card5, card6 = st.columns([1.5, 1, 1, 1, 1, 1])
    card1.metric("Simulation Clock", current_time.strftime("%Y-%m-%d %H:%M"))
    card2.metric("Current Request Step", f"{st.session_state.playback_frame + 1}/{max_step + 1}")
    card3.metric("Processed Decisions", frame_data["processed_count"])
    card4.metric("Assigned So Far", frame_data["assigned_count"])
    card5.metric("Pending Ready", len(pending_queue_rows))
    card6.metric("Not Ready", len(not_ready_waiting_rows))

    routing_event_label = str(current_event.get("event_type", "")).replace("_", " ").title()
    routing_detail_label = str(current_event.get("details", "")).replace("_", " ").title()
    st.markdown(
        "**Current Routing Decision:** "
        f"{routing_event_label} | "
        f"Request={current_event.get('request_id')} | "
        f"Staff={format_staff_label(current_event.get('staff_id'), staff_college_map)} | "
        f"Details={routing_detail_label}"
    )

    queue_col, waiting_col = st.columns(2)
    with queue_col:
        st.subheader("Pending Queue")
        st.caption("Ready requests that have arrived but cannot be assigned at this step.")
        if pending_queue_rows:
            render_theme_table(pd.DataFrame(pending_queue_rows), height_px=320)
        else:
            st.caption("No ready requests are blocked from assignment at this step.")

    with waiting_col:
        st.subheader("Unassignable Waiting List")
        st.caption("Arrived requests still blocked by requirements or payment readiness.")
        if not_ready_waiting_rows:
            render_theme_table(pd.DataFrame(not_ready_waiting_rows), height_px=320)
        else:
            st.caption("No not-ready requests at this step.")

    st.subheader("Staff Capacity View")
    quota_enforced = results.get("allocator_type") != "quota_free"
    current_day = current_time.date()
    capacity_rows = []
    for staff in engine.staff_pool:
        rows_for_staff = staff_rows.get(staff.staff_id, [])
        assigned_today = sum(1 for row in rows_for_staff if row.get("_date") == current_day)
        total_assigned = len(rows_for_staff)
        row = {
            "Staff ID": staff.staff_id,
            "Staff": format_staff_label(staff.staff_id, staff_college_map),
            "College": staff.college_affiliation,
            "Assigned Today": assigned_today,
            "Total Assigned": total_assigned,
        }
        if quota_enforced:
            row["Quota/Day"] = staff.quota_limit
            row["Today Fill %"] = round((assigned_today / max(staff.quota_limit, 1)) * 100.0, 1)
        capacity_rows.append(row)

    capacity_df = pd.DataFrame(capacity_rows)
    assigned_today_map = {row["Staff ID"]: row["Assigned Today"] for row in capacity_rows}
    total_assigned_map = {row["Staff ID"]: row["Total Assigned"] for row in capacity_rows}

    fig_capacity = go.Figure()
    fig_capacity.add_trace(
        go.Bar(
            name="Assigned Today",
            x=capacity_df["Staff"],
            y=capacity_df["Assigned Today"],
            marker_color="#a855f7",
            text=capacity_df["Assigned Today"],
            textposition="outside",
        )
    )
    if quota_enforced:
        fig_capacity.add_trace(
            go.Bar(
                name="Quota per Day",
                x=capacity_df["Staff"],
                y=capacity_df["Quota/Day"],
                marker_color="#22d3ee",
                opacity=0.55,
                text=capacity_df["Quota/Day"],
                textposition="outside",
            )
        )
    fig_capacity.update_layout(
        title="Daily Staff Capacity at Current Request Step",
        xaxis_title="Staff",
        yaxis_title="Requests",
        barmode="group",
        height=360,
    )
    apply_plot_theme(fig_capacity)
    st.plotly_chart(fig_capacity, use_container_width=True, config={"staticPlot": True})

    st.subheader("Live Staff Request Lists")
    ordered_staff_ids = [staff.staff_id for staff in engine.staff_pool]
    for idx in range(0, len(ordered_staff_ids), 3):
        row_ids = ordered_staff_ids[idx: idx + 3]
        row_cols = st.columns(3)
        for col, staff_id in zip(row_cols, row_ids):
            with col:
                meta = staff_meta.get(staff_id, {"college": "-", "quota": None})
                assigned_today = assigned_today_map.get(staff_id, 0)
                total_assigned = total_assigned_map.get(staff_id, 0)
                if quota_enforced and meta.get("quota") is not None:
                    quota_value = int(meta["quota"])
                    st.metric(f"{staff_id} ({meta['college']})", f"{assigned_today}/{quota_value}", delta=f"Total {total_assigned}")
                    if assigned_today >= quota_value:
                        st.warning("Quota full at this step")
                else:
                    st.metric(f"{staff_id} ({meta['college']})", f"{total_assigned} assigned")

                rows = staff_rows.get(staff_id, [])
                if rows:
                    render_theme_table(pd.DataFrame(staff_rows_with_day_separators(rows)), height_px=340)
                else:
                    st.caption("No requests routed here yet.")

    if st.session_state.playback_playing:
        if st.session_state.playback_frame < max_step:
            tm.sleep(SPEED_OPTIONS.get(st.session_state.playback_speed, 0.45))
            st.session_state.playback_frame += 1
            st.rerun(scope="fragment")
        else:
            st.session_state.playback_playing = False


render_playback_section(results, engine)


# ============================================================================
# KPI METRICS
# ============================================================================
st.divider()


st.header("Key Metrics")

custom_m = results.get("custom_metrics", {"avg_waiting_time_hours": 0.0, "avg_turnaround_days": 0.0, "total_processed": 0})
gen_m = results.get("generated_metrics", {"avg_waiting_time_hours": 0.0, "avg_turnaround_days": 0.0, "total_processed": 0})

custom_count = custom_m.get("total_processed", 0)
gen_count = gen_m.get("total_processed", 0)

all_requests = results.get("generated_requests", [])

submitted_custom = sum(1 for req in all_requests if req.get("is_custom"))
submitted_generated = sum(1 for req in all_requests if not req.get("is_custom"))

k1, k2, k3, k4, k5 = st.columns(5)

with k1:
    processed = gen_count + custom_count
    expected = submitted_generated + submitted_custom

    pct = (processed / expected * 100.0) if expected > 0 else 0.0

    if submitted_custom > 0:
        delta_str = f"{gen_count} Gen | {custom_count}/{submitted_custom} Cust"
    else:
        delta_str = f"{pct:.0f}%"

    if st.session_state.get("disable_generated_requests", False):
        st.metric("Total Processed", f"{processed}/{expected} Custom", delta_str)
    else:
        st.metric("Total Processed", f"{processed}/{expected}", delta_str)
with k2:
    avg_wait_hours = float(results.get("avg_waiting_time_hours", 0.0))
    if custom_count > 0 and gen_count > 0:
        delta_wait = f"Gen: {gen_m.get('avg_waiting_time_hours', 0.0):.1f}h | Cust: {custom_m.get('avg_waiting_time_hours', 0.0):.1f}h"
    else:
        delta_wait = f"{(avg_wait_hours / 24.0):.2f} d equiv"
    st.metric("Avg Queue Wait", f"{avg_wait_hours:.2f} h", delta_wait)
with k3:
    avg_turn_days = float(results.get("avg_turnaround_days", 0.0))
    if custom_count > 0 and gen_count > 0:
        delta_turn = f"Gen: {gen_m.get('avg_turnaround_days', 0.0):.1f}d | Cust: {custom_m.get('avg_turnaround_days', 0.0):.1f}d"
    else:
        delta_turn = f"{(avg_turn_days * 24.0):.2f} h equiv"
    st.metric("Avg Turnaround", f"{avg_turn_days:.2f} d", delta_turn)
with k4:
    elapsed_days = float(results.get("total_days_elapsed", 0.0))
    elapsed_day_index = max(1, int(math.floor(elapsed_days)) + 1)
    st.metric(
        "Days Elapsed",
        f"Day {elapsed_day_index}",
        f"{elapsed_days:.2f} d span",
    )
with k5:
    st.metric("Throughput", f"{results.get('throughput_req_per_day', 0):.2f} req/day")

if results.get("absent_staff"):
    st.warning("Absent staff: " + ", ".join(results.get("absent_staff", [])))

variant_label = format_variant_label(
    results.get('scheduler_type', ''), results.get('allocator_type', '')
)

st.header("Visualizations")

st.markdown(
    f"**Scheduler:** {humanize_option_label(results.get('scheduler_type'))}<br>"
    f"**Allocator:** {humanize_option_label(results.get('allocator_type'))}"
    , unsafe_allow_html=True,
)

labels_outside_default = st.session_state.get("labels_outside", True)
st.checkbox(
    "Show labels outside charts (useful for filtering)",
    value=labels_outside_default,
    key="labels_outside",
)

fig_41 = build_baseline_queue_dynamics_chart(results.get("event_log", []), variant_label)
if fig_41.data:
    st.plotly_chart(fig_41, use_container_width=True)
else:
    st.info("Queue dynamics are not available for the selected simulation.")

if results.get("scheduler_type") == "WEIGHTED":
    generated_requests = results.get("generated_requests", [])
    request_types = sorted(
        {
            (req.get("document_type") if isinstance(req, dict) else getattr(req, "document_type", None))
            for req in generated_requests
            if req is not None
        }
    )
    request_types = [rt for rt in request_types if rt]
    selected_request_types = st.multiselect(
        "Request types to include in the priority distribution",
        options=request_types,
        default=request_types,
        key="priority_distribution_doc_types",
    )

    if not selected_request_types:
        st.info("Select one or more request types to display the priority score distribution.")
    else:
        fig_43 = build_weighted_priority_distribution_chart(
            generated_requests,
            selected_request_types,
            variant_label,
        )
        if fig_43.data:
            st.plotly_chart(fig_43, use_container_width=True)
        else:
            st.info("Priority score distribution is not available for the selected simulation.")
else:
    st.info("Priority score distribution is only shown for the WEIGHTED scheduler.")

# ============================================================================
# STAFF LOAD + TIMELINE CHARTS
# ============================================================================

st.header("Staff and Timeline")

staff_load = results.get("staff_load", {})
if staff_load:
    staff_ids = list(staff_load.keys())
    staff_labels = [format_staff_label(staff_id, staff_college_map) for staff_id in staff_ids]
    staff_values = [staff_load[staff_id] for staff_id in staff_ids]
    fig_staff = go.Figure(
        data=[
            go.Bar(
                x=staff_labels,
                y=staff_values,
                marker=dict(
                    color=staff_values,
                    colorscale=[
                        [0.0, "#5b21b6"],
                        [0.55, "#9333ea"],
                        [1.0, "#22d3ee"],
                    ],
                ),
                text=staff_values,
                textposition="outside",
            )
        ]
    )
    fig_staff.update_layout(
        title=f"Requests Processed per Staff — {variant_label}",
        xaxis_title="Staff",
        yaxis_title="Processed Requests",
        height=450,
    )
    apply_plot_theme(fig_staff)
    st.plotly_chart(fig_staff, use_container_width=True)

if engine.completed:
    timeline_rows = []
    for req in engine.completed:
        assigned_day = (req.assignment_time.date() - engine.start_time.date()).days + 1
        timeline_rows.append(
            {
                "Assigned Day": assigned_day,
                "College": req.college,
                "Count": 1,
            }
        )
    timeline_df = pd.DataFrame(timeline_rows)
    grouped = timeline_df.groupby(["Assigned Day", "College"]).size().reset_index(name="Count")
    fig_timeline = px.bar(
        grouped,
        x="Assigned Day",
        y="Count",
        color="College",
        color_discrete_sequence=CHART_COLORWAY,
        title=f"Assignments per Day by College — {variant_label}",
        height=500,
        text="Count" if st.session_state.labels_outside else None,
    )
    apply_plot_theme(fig_timeline)
    st.plotly_chart(fig_timeline, use_container_width=True)

if engine.completed:
    selected_college = st.selectbox(
        "Filter Document Mix by College",
        ["All"] + COLLEGES,
        key="doc_mix_college",
    )
    if selected_college == "All":
        filtered_docs = list(engine.completed)
    else:
        filtered_docs = [
            req for req in engine.completed if req.college == selected_college
        ]

    if filtered_docs:
        doc_df = pd.DataFrame(
            [{"document_type": req.document_type} for req in filtered_docs]
        )
        doc_counts = doc_df["document_type"].value_counts().reset_index()
        doc_counts.columns = ["Document", "Count"]
        total_docs = int(doc_counts["Count"].sum())

        short_doc_names = {
            "Certification, Authentication and Verification (CAV)": "CAV",
            "Official Transcript of Records (TOR) and Transfer Credentials (TC)": "TOR/TC",
            "Evaluation of Grades; Report of Grades (ROG); Certificate of Registration (COR)": "ROG/COR",
            "Permit to Cross-Enrol": "Cross-Enrol Permit",
            "Academic Load Revision (ALRP)": "ALRP",
            "Shifter’s Form, Returnee’s Form or Leave of Absence": "Shifter/Returnee/LOA",
            "Registration of Old and Returnee Students": "Old/Returnee Reg",
        }

        doc_left, doc_right = st.columns([2, 1])
        with doc_left:
            fig_docs = go.Figure(
                data=[
                    go.Pie(
                        labels=doc_counts["Document"],
                        values=doc_counts["Count"],
                        hole=0.55,
                        textinfo="percent",
                    )
                ]
            )
            fig_docs.update_layout(
                title=f"Requested Document Mix (Completed) — {variant_label}",
                height=420,
                showlegend=False,
            )
            apply_plot_theme(fig_docs)
            st.plotly_chart(fig_docs, use_container_width=True)

        with doc_right:
            st.markdown(
                f"<div style='font-size:0.9rem; color:#b6b0d4; margin-bottom:0.4rem;'>"
                f"Total processed: <span style='color:#f5f3ff; font-weight:700;'>{total_docs}</span>"
                f"</div>",
                unsafe_allow_html=True,
            )
            legend_rows = []
            for row in doc_counts.itertuples(index=False):
                pct = (float(row.Count) / max(total_docs, 1)) * 100.0
                label = short_doc_names.get(row.Document, row.Document)
                legend_rows.append(
                    {
                        "label": label,
                        "count": int(row.Count),
                        "pct": pct,
                    }
                )
            mid_point = (len(legend_rows) + 1) // 2
            legend_col1, legend_col2 = st.columns(2)
            for col, rows in zip(
                [legend_col1, legend_col2],
                [legend_rows[:mid_point], legend_rows[mid_point:]],
            ):
                with col:
                    block = []
                    for item in rows:
                        block.append(
                            f"<div style='margin:0.3rem 0;'>"
                            f"<span style='font-weight:700; color:#e7e1fa;'>{item['label']}</span>"
                            f"<br />"
                            f"<span style='color:#a7a0c5;'>{item['count']}"
                            f" <span style='color:#7dd3fc;'>({item['pct']:.1f}%)</span></span>"
                            f"</div>"
                        )
                    st.markdown("".join(block), unsafe_allow_html=True)
    else:
        st.caption("No completed requests for the selected college.")

# ============================================================================
# REQUEST INSPECTION
# ============================================================================

st.header("Request Inspection")

completed_requests = engine.completed

if not completed_requests:
    st.info("No completed requests to inspect.")
else:
    f1, f2, f3 = st.columns(3)

    with f1:
        filter_college = st.selectbox("Filter by College", ["All"] + COLLEGES)
    with f2:
        filter_doc = st.selectbox("Filter by Document", ["All"] + list(DOCUMENT_COMPLEXITY.keys()))
    with f3:
        sort_options = [
            "Assigned Day",
            "Submission Time",
            "Staff List Order",
            "Queue Wait Desc",
            "Queue Wait Asc",
            "Turnaround Desc",
        ]
        if is_weighted_scheduler:
            sort_options = ["Priority Desc", "Priority Asc"] + sort_options

        sort_by = st.selectbox(
            "Sort by",
            sort_options,
            index=0,
        )

    filtered = completed_requests
    if filter_college != "All":
        filtered = [req for req in filtered if req.college == filter_college]
    if filter_doc != "All":
        filtered = [req for req in filtered if req.document_type == filter_doc]

    staff_order = {staff.staff_id: idx for idx, staff in enumerate(engine.staff_pool)}

    if sort_by == "Priority Desc":
        filtered = sorted(
            filtered,
            key=lambda req: (-float(getattr(req, "priority_score", 0.0)), req.submission_time),
        )
    elif sort_by == "Priority Asc":
        filtered = sorted(
            filtered,
            key=lambda req: (float(getattr(req, "priority_score", 0.0)), req.submission_time),
        )
    elif sort_by == "Assigned Day":
        filtered = sorted(
            filtered,
            key=lambda req: (
                req.assignment_time.date(),
                req.completion_time,
                req.submission_time,
            ),
        )
    elif sort_by == "Submission Time":
        filtered = sorted(filtered, key=lambda req: req.submission_time)
    elif sort_by == "Staff List Order":
        filtered = sorted(
            filtered,
            key=lambda req: (
                req.assignment_time,
                staff_order.get(req.assigned_staff, 9999),
                req.submission_time,
            ),
        )
    elif sort_by == "Queue Wait Desc":
        filtered = sorted(filtered, key=lambda req: req.get_waiting_time_minutes(), reverse=True)
    elif sort_by == "Queue Wait Asc":
        filtered = sorted(filtered, key=lambda req: req.get_waiting_time_minutes())
    else:
        filtered = sorted(filtered, key=lambda req: req.get_turnaround_time_minutes(), reverse=True)

    table_rows = []
    for idx, req in enumerate(filtered):
        assigned_day = (req.assignment_time.date() - engine.start_time.date()).days + 1
        table_rows.append(
            {
                "Row": idx + 1,
                "Request": f"⭐ {req.request_id}" if getattr(req, "is_custom", False) else req.request_id,
                "College": req.college,
                "Document": req.document_type,
                "Completeness": round(float(getattr(req, "completeness_of_requirements", 0.0)), 2),
                "Requester Status": getattr(req, "requester_type", "-"),
                "Payment Status": getattr(req, "payment_status", "-"),
                "Priority Score": round(float(getattr(req, "priority_score", 0.0)), 4),
                "Queue Wait (h)": round(req.get_waiting_time_minutes() / 60.0, 2),
                "Turnaround (d)": round(req.get_turnaround_time_minutes() / 1440.0, 2),
                "Assigned Day": assigned_day,
                "Staff": format_staff_label(req.assigned_staff, staff_college_map),
            }
        )

    table_df = pd.DataFrame(table_rows)
    render_theme_table(table_df, height_px=430)

    st.subheader("Detailed Request Panel")
    pick_index = st.number_input(
        "Select Row",
        min_value=1,
        max_value=max(1, len(filtered)),
        value=1,
        step=1,
    )

    selected_req = filtered[int(pick_index) - 1]
    d1, d2 = st.columns(2)

    with d1:
        st.write(f"**Request ID:** {selected_req.request_id}")
        st.write(f"**College:** {selected_req.college}")
        st.write(f"**Document Type:** {selected_req.document_type}")
        st.write(
            f"**Completeness:** {float(getattr(selected_req, 'completeness_of_requirements', 0.0)):.2f}"
        )
        st.write(f"**Requester Status:** {getattr(selected_req, 'requester_type', '-')}")
        st.write(f"**Urgency (generated):** {getattr(selected_req, 'urgency', '-')}")
        st.write(f"**Payment Status:** {getattr(selected_req, 'payment_status', '-')}")
        st.write(f"**Priority Score:** {float(getattr(selected_req, 'priority_score', 0.0)):.4f}")
        st.write(f"**Assigned Staff:** {format_staff_label(selected_req.assigned_staff, staff_college_map)}")
        for field in CUSTOM_REQUEST_FIELDS:
            field_key = str(field.get("key"))
            if not field_key:
                continue
            extra_fields = getattr(selected_req, "extra_fields", {}) or {}
            if field_key in extra_fields:
                score_value = extra_fields.get(f"{field_key}_score")
                score_text = f" ({float(score_value):.2f})" if score_value is not None else ""
                st.write(f"**{field.get('label', field_key)}:** {extra_fields.get(field_key)}{score_text}")

    with d2:
        st.write(f"**Submission:** {format_compact_datetime(selected_req.submission_time)}")
        st.write(
            "**Requirements Partial:** "
            f"{format_compact_datetime(getattr(selected_req, 'requirements_partial_time', None))}"
        )
        st.write(
            "**Requirements Complete:** "
            f"{format_compact_datetime(getattr(selected_req, 'requirements_complete_time', None))}"
        )
        st.write(
            "**Payment Time:** "
            f"{format_compact_datetime(getattr(selected_req, 'payment_time', None))}"
        )
        st.write(
            "**Ready Time:** "
            f"{format_compact_datetime(getattr(selected_req, 'ready_time', None))}"
        )
        st.write(f"**Assignment:** {format_compact_datetime(selected_req.assignment_time)}")
        st.write(f"**Completion:** {format_compact_datetime(selected_req.completion_time)}")
        st.write(f"**Queue Wait:** {selected_req.get_waiting_time_minutes() / 60.0:.2f} h")
        st.write(f"**Turnaround:** {selected_req.get_turnaround_time_minutes() / 1440.0:.2f} d")

    if is_weighted_scheduler:
        st.subheader("Priority Score Progression")

        def _score_at(request_obj: DocumentRequest, at_time: Optional[datetime]) -> Optional[float]:
            if at_time is None:
                return None
            original_state = (
                request_obj.completeness_of_requirements,
                request_obj.payment_status,
                request_obj.requirements_stage,
                request_obj.priority_score,
            )
            try:
                try:
                    return request_obj.calculate_priority(
                        at_time,
                        engine.priority_weights,
                        engine.workday_minutes,
                        urgency=engine.urgency,
                        custom_criteria=getattr(engine, "custom_criteria", []),
                    )
                except TypeError as exc:
                    if "custom_criteria" not in str(exc):
                        raise
                    return request_obj.calculate_priority(
                        at_time,
                        engine.priority_weights,
                        engine.workday_minutes,
                        urgency=engine.urgency,
                    )
            finally:
                (
                    request_obj.completeness_of_requirements,
                    request_obj.payment_status,
                    request_obj.requirements_stage,
                    request_obj.priority_score,
                ) = original_state

        stage_points = [
            ("Submitted", selected_req.submission_time),
            ("Requirements Partial", getattr(selected_req, "requirements_partial_time", None)),
            ("Requirements Complete", getattr(selected_req, "requirements_complete_time", None)),
            ("Payment", getattr(selected_req, "payment_time", None)),
            ("Ready", getattr(selected_req, "ready_time", None)),
            ("Assigned", selected_req.assignment_time),
        ]
        stage_rows = []
        for label, ts in stage_points:
            if ts is None:
                continue
            score_value = _score_at(selected_req, ts)
            stage_rows.append(
                {
                    "Stage": label,
                    "Time": format_compact_datetime(ts),
                    "Priority Score": round(float(score_value or 0.0), 4),
                }
            )
        if stage_rows:
            stage_df = pd.DataFrame(stage_rows)
            render_theme_table(stage_df, height_px=220)
            with st.expander("Debug: weight & contribution breakdown", expanded=False):
                st.write("Engine priority_weights:")
                st.write(dict(sorted(
                    engine.priority_weights.items(),
                    key=lambda item: (-float(item[1]), format_criterion_label(item[0])),
                )))
        else:
            st.write("No stage timestamps available.")

    st.subheader("Request Lifecycle Timeline")

    lifecycle_times = [
        selected_req.submission_time,
        selected_req.assignment_time,
        selected_req.completion_time,
    ]
    lifecycle_labels = [
        "Submitted",
        f"Assigned ({format_staff_label(selected_req.assigned_staff, staff_college_map)})",
        "Completed",
    ]
    lifecycle_colors = ["#22d3ee", "#a855f7", "#f472b6"]
    lifecycle_y = [2.0, 1.0, 0.0]

    step_x = [
        lifecycle_times[0],
        lifecycle_times[1],
        lifecycle_times[1],
        lifecycle_times[2],
        lifecycle_times[2],
    ]
    step_y = [
        lifecycle_y[0],
        lifecycle_y[0],
        lifecycle_y[1],
        lifecycle_y[1],
        lifecycle_y[2],
    ]

    fig_request_timeline = go.Figure()
    fig_request_timeline.add_trace(
        go.Scatter(
            x=step_x,
            y=step_y,
            mode="lines",
            line=dict(color="#8b79bb", width=3),
            hoverinfo="skip",
            showlegend=False,
        )
    )
    fig_request_timeline.add_trace(
        go.Scatter(
            x=lifecycle_times,
            y=lifecycle_y,
            mode="markers+text",
            text=lifecycle_labels,
            textposition=["top left", "middle right", "bottom right"],
            marker=dict(size=14, color=lifecycle_colors),
            showlegend=False,
            customdata=lifecycle_labels,
            hovertemplate="%{customdata}<br>%{x}<extra></extra>",
        )
    )
    fig_request_timeline.update_layout(
        height=320,
        xaxis_title="Time",
        yaxis=dict(
            showticklabels=False,
            showgrid=False,
            zeroline=False,
            range=[-0.5, 2.5],
        ),
        margin=dict(l=20, r=20, t=20, b=20),
    )
    apply_plot_theme(fig_request_timeline)
    fig_request_timeline.update_yaxes(showgrid=False)
    st.plotly_chart(fig_request_timeline, use_container_width=True)


# ============================================================================
# COMPARISON TOOLS
# ============================================================================
st.divider()

st.header("Comparison Tools")

c1, c2 = st.columns(2)
with c1:
    compare_schedulers = st.multiselect(
        "Schedulers",
        SCHEDULER_OPTIONS,
        default=SCHEDULER_OPTIONS,
        format_func=lambda value: SCHEDULER_LABELS.get(value, value),
    )
with c2:
    compare_allocators = st.multiselect(
        "Allocators",
        ALLOCATOR_OPTIONS,
        default=ALLOCATOR_OPTIONS,
        format_func=lambda value: ALLOCATOR_LABELS.get(value, value.replace("_", " ").title()),
    )

if st.button("Run Comparison Across Selected Variants", use_container_width=True):
    if not compare_schedulers or not compare_allocators:
        st.warning("Select at least one scheduler and one allocator.")
    else:
        compare_rows = []
        compare_details = []
        same_seed = int(results.get("seed_used", st.session_state.manual_seed))

        base_payload = build_api_payload()
        
        with st.spinner("Running comparison simulations on backend..."):
            for scheduler in compare_schedulers:
                for allocator in compare_allocators:
                    payload = base_payload.copy()
                    payload["scheduler_type"] = scheduler
                    payload["allocator_type"] = allocator
                    payload["random_seed"] = same_seed
                    
                    try:
                        response = requests.post(f"{BACKEND_URL}/simulate", json=payload, timeout=120)
                        if response.status_code == 200:
                            data = response.json()
                            compare_result = data.get("results", {})
                            
                            compare_details.append(
                                {
                                    "scheduler": scheduler,
                                    "allocator": allocator,
                                    "completed_requests": compare_result.get("completed_requests", []),
                                }
                            )
                            staff_load_values = list(compare_result.get("staff_load", {}).values())
                            if not staff_load_values:
                                staff_load_std = 0.0
                                staff_load_mean = 0.0
                                staff_load_cv = 0.0
                            else:
                                s = pd.Series(staff_load_values)
                                staff_load_std = float(s.std(ddof=0))
                                staff_load_mean = float(s.mean())
                                staff_load_cv = round(staff_load_std / staff_load_mean, 4) if staff_load_mean > 0 else 0.0
                            
                            compare_rows.append(
                                {
                                    "scheduler": scheduler,
                                    "allocator": allocator,
                                    "total_processed": compare_result.get("total_processed", 0),
                                    "avg_waiting_time_hours": compare_result.get("avg_waiting_time_hours", 0.0),
                                    "avg_turnaround_days": compare_result.get("avg_turnaround_days", 0.0),
                                    "total_days_elapsed": compare_result.get("total_days_elapsed", 0.0),
                                    "throughput_req_per_day": compare_result.get("throughput_req_per_day", 0.0),
                                    "staff_load_std": round(staff_load_std, 2),
                                    "staff_load_cv": round(staff_load_cv, 4),
                                }
                            )
                    except requests.exceptions.RequestException as e:
                        st.error(f"Failed to run {scheduler} + {allocator}: {e}")
        
        compare_df = pd.DataFrame(compare_rows)
        baseline = compare_df[
            (compare_df["scheduler"] == "FCFS")
            & (compare_df["allocator"] == "college_based")
        ]
        if baseline.empty:
            baseline_row = compare_df.iloc[0]
        else:
            baseline_row = baseline.iloc[0]
        compare_df["delta_wait_vs_baseline"] = (
            compare_df["avg_waiting_time_hours"] - baseline_row["avg_waiting_time_hours"]
        ).round(2)
        compare_df["delta_throughput_vs_baseline"] = (
            compare_df["throughput_req_per_day"] - baseline_row["throughput_req_per_day"]
        ).round(2)
        compare_df["delta_turnaround_vs_baseline"] = (
            compare_df["avg_turnaround_days"] - baseline_row["avg_turnaround_days"]
        ).round(2)
        st.session_state.comparison_df = compare_df
        st.session_state.comparison_details = compare_details


if st.session_state.comparison_df is not None:
    compare_df = st.session_state.comparison_df.copy()

    ALLOCATOR_LABELS = {
        "college_based": "College Based",
        "workload_based": "Workload Based",
        "pooled": "Pooled",
        "quota_free": "Quota Free",
    }

    scheduler_display = compare_df["scheduler"].map(lambda s: SCHEDULER_LABELS.get(s, str(s)))
    allocator_display = compare_df["allocator"].map(lambda a: ALLOCATOR_LABELS.get(a, str(a).replace("_", " ").title()))
    compare_df["Variant"] = scheduler_display + " | " + allocator_display

    st.subheader("Variant Comparison Table")
    st.caption("Use the filters below to select which variants and columns appear in the comparison table.")

    variant_options = list(compare_df["Variant"].unique())
    selected_variants = st.multiselect("Show variants", variant_options, default=variant_options)
    if selected_variants:
        compare_df = compare_df[compare_df["Variant"].isin(selected_variants)].copy()

    available_cols = [
        c
        for c in compare_df.columns
        if c not in ("Variant", "scheduler", "allocator")
    ]
    default_cols = ["Variant"] + available_cols
    selected_cols = st.multiselect(
        "Columns to display",
        options=default_cols,
        default=default_cols,
        format_func=humanize_option_label,
    )

    if not selected_cols:
        st.info("Select at least one column to display the comparison table.")
    else:
        def human_label(col: str) -> str:
            labels = {
                "scheduler": "Scheduler",
                "allocator": "Allocator",
                "Variant": "Variant",
                "total_processed": "Total Processed",
                "avg_waiting_time_hours": "Avg Waiting Time (h)",
                "avg_turnaround_days": "Avg Turnaround (d)",
                "total_days_elapsed": "Total Days Elapsed",
                "throughput_req_per_day": "Throughput (req/day)",
                "staff_load_std": "Staff Load Std Dev",
                "staff_load_cv": "Staff Load CV",
                "delta_wait_vs_baseline": "Δ Wait vs Baseline (h)",
                "delta_throughput_vs_baseline": "Δ Throughput vs Baseline",
                "delta_turnaround_vs_baseline": "Δ Turnaround vs Baseline (d)",
                "order_changed_pct": "Order Changed (%)",
            }
            if col in labels:
                return labels[col]
            # fallback: turn snake_case into Title Case
            return str(col).replace("_", " ").title()

        # Build display DataFrame and rename columns for presentation
        display_df = compare_df[selected_cols].copy()
        rename_map = {c: human_label(c) for c in display_df.columns}
        display_df.rename(columns=rename_map, inplace=True)

        render_theme_table(display_df, height_px=420)

        # Also show charts for the filtered variants
        imbalance_fig = build_workload_imbalance_chart(compare_df)
        if imbalance_fig.data:
            st.subheader("Workload Imbalance by Variant")
            st.plotly_chart(imbalance_fig, use_container_width=True)

        summary_fig = build_variant_summary_chart(compare_df)
        if summary_fig.data:
            st.subheader("Variant Performance Summary")
            st.plotly_chart(summary_fig, use_container_width=True)

comparison_details = st.session_state.get("comparison_details")
if st.session_state.comparison_df is not None and comparison_details:
    st.subheader("Request-Level Differences")

    def _build_request_index(completed_requests: List[Dict]) -> Dict[str, Dict[str, object]]:
        rows = []
        for item in completed_requests:
            request_id = item.get("request_id")
            if not request_id:
                continue
            assign_raw = item.get("assignment_time")
            complete_raw = item.get("completion_time")
            rows.append(
                {
                    "request_id": request_id,
                    "assignment_time": parse_event_time(str(assign_raw)) if assign_raw else None,
                    "completion_time": parse_event_time(str(complete_raw)) if complete_raw else None,
                    "assigned_staff": item.get("assigned_staff"),
                }
            )

        rows = sorted(
            rows,
            key=lambda r: (r["assignment_time"] or datetime.max, r["request_id"]),
        )
        index: Dict[str, Dict[str, object]] = {}
        for rank, row in enumerate(rows, start=1):
            index[row["request_id"]] = {
                "rank": rank,
                "assignment_time": row["assignment_time"],
                "completion_time": row["completion_time"],
                "assigned_staff": row["assigned_staff"],
            }
        return index

    compare_df = st.session_state.comparison_df
    baseline_row = compare_df.iloc[0]
    baseline_match = compare_df[
        (compare_df["scheduler"] == "FCFS")
        & (compare_df["allocator"] == "college_based")
    ]
    if not baseline_match.empty:
        baseline_row = baseline_match.iloc[0]

    baseline_key = (baseline_row["scheduler"], baseline_row["allocator"])
    baseline_details = next(
        (
            item
            for item in comparison_details
            if (item["scheduler"], item["allocator"]) == baseline_key
        ),
        None,
    )

    if baseline_details is None:
        st.info("Baseline details not available for request-level comparison.")
    else:
        baseline_index = _build_request_index(baseline_details["completed_requests"])
        baseline_requests = set(baseline_index.keys())

        diff_rows = []
        for item in comparison_details:
            scheduler = item["scheduler"]
            allocator = item["allocator"]
            key = (scheduler, allocator)
            if key == baseline_key:
                continue

            current_index = _build_request_index(item["completed_requests"])
            current_requests = set(current_index.keys())
            common = baseline_requests.intersection(current_requests)

            if not common:
                continue

            order_changed = 0
            staff_changed = 0
            rank_shift_total = 0.0
            assign_delta_total = 0.0
            complete_delta_total = 0.0

            for request_id in common:
                base = baseline_index[request_id]
                current = current_index[request_id]
                if base["rank"] != current["rank"]:
                    order_changed += 1
                    rank_shift_total += abs(current["rank"] - base["rank"])
                if base.get("assigned_staff") != current.get("assigned_staff"):
                    staff_changed += 1

                base_assign = base.get("assignment_time")
                current_assign = current.get("assignment_time")
                if base_assign and current_assign:
                    assign_delta_total += abs((current_assign - base_assign).total_seconds()) / 60.0

                base_complete = base.get("completion_time")
                current_complete = current.get("completion_time")
                if base_complete and current_complete:
                    complete_delta_total += abs((current_complete - base_complete).total_seconds()) / 60.0

            total_common = len(common)
            avg_rank_shift = rank_shift_total / max(order_changed, 1)
            avg_assign_delta = assign_delta_total / total_common
            avg_complete_delta = complete_delta_total / total_common

            diff_rows.append(
                {
                    "scheduler": scheduler,
                    "allocator": allocator,
                    "order_changed_count": order_changed,
                    "order_changed_pct": round((order_changed / total_common) * 100.0, 2),
                    "avg_abs_rank_shift": round(avg_rank_shift, 2),
                    "staff_changed_count": staff_changed,
                    "staff_changed_pct": round((staff_changed / total_common) * 100.0, 2),
                    "avg_assign_time_delta_min": round(avg_assign_delta, 2),
                    "avg_complete_time_delta_min": round(avg_complete_delta, 2),
                }
            )

        diff_df = pd.DataFrame(diff_rows)
        if diff_df.empty:
            st.info("No comparable request-level differences found.")
        else:
            raw_variant_options = [
                (row["scheduler"], row["allocator"]) for row in diff_rows
            ]

            diff_df = diff_df.rename(
                columns={
                    "scheduler": "Scheduler",
                    "allocator": "Allocator",
                    "order_changed_count": "Order Changed Count",
                    "order_changed_pct": "Order Changed %",
                    "avg_abs_rank_shift": "Avg Abs Rank Shift",
                    "staff_changed_count": "Staff Changed Count",
                    "staff_changed_pct": "Staff Changed %",
                    "avg_assign_time_delta_min": "Avg Assign Delta (min)",
                    "avg_complete_time_delta_min": "Avg Complete Delta (min)",
                }
            )
            diff_df["Scheduler"] = diff_df["Scheduler"].map(
                lambda s: SCHEDULER_LABELS.get(s, str(s))
            )
            diff_df["Allocator"] = diff_df["Allocator"].map(
                lambda a: ALLOCATOR_LABELS.get(a, str(a).replace("_", " ").title())
            )

            render_theme_table(diff_df, height_px=320)

            fig_diff = go.Figure()
            fig_diff.add_trace(
                go.Bar(
                    name="Order Changed %",
                    x=diff_df["Allocator"],
                    y=diff_df["Order Changed %"],
                    marker_color="#a855f7",
                    text=diff_df["Order Changed %"].apply(lambda v: f"{v:.1f}%"),
                    textposition="outside",
                )
            )
            fig_diff.add_trace(
                go.Bar(
                    name="Staff Changed %",
                    x=diff_df["Allocator"],
                    y=diff_df["Staff Changed %"],
                    marker_color="#22d3ee",
                    text=diff_df["Staff Changed %"].apply(lambda v: f"{v:.1f}%"),
                )
            )
            fig_diff.update_layout(
                title="Request-Level Changes vs Baseline",
                xaxis_title="Allocator",
                yaxis_title="Percent of Requests",
                barmode="group",
                height=320,
            )
            apply_plot_theme(fig_diff)
            st.plotly_chart(fig_diff, use_container_width=True)

            selected_variant = st.selectbox(
                "Inspect Variant",
                options=raw_variant_options,
                format_func=lambda v: f"{SCHEDULER_LABELS.get(v[0], v[0])} | {ALLOCATOR_LABELS.get(v[1], v[1].replace('_', ' ').title())}",
                index=0,
            )

            if selected_variant:
                selected_sched, selected_alloc = selected_variant
                selected_detail = next(
                    (
                        item
                        for item in comparison_details
                        if item["scheduler"] == selected_sched
                        and item["allocator"] == selected_alloc
                    ),
                    None,
                )
                if selected_detail:
                    current_index = _build_request_index(
                        selected_detail["completed_requests"]
                    )
                    change_rows = []
                    for request_id in baseline_requests.intersection(current_index.keys()):
                        base = baseline_index[request_id]
                        current = current_index[request_id]
                        rank_shift = current["rank"] - base["rank"]
                        base_assign = base.get("assignment_time")
                        current_assign = current.get("assignment_time")
                        assign_delta = None
                        if base_assign and current_assign:
                            assign_delta = round(
                                (current_assign - base_assign).total_seconds() / 60.0, 2
                            )
                        change_rows.append(
                            {
                                "Request": request_id,
                                "Rank Shift": rank_shift,
                                "Assigned Staff": current.get("assigned_staff"),
                                "Staff Changed": base.get("assigned_staff")
                                != current.get("assigned_staff"),
                                "Assign Delta (min)": assign_delta,
                            }
                        )

                    change_df = pd.DataFrame(change_rows)
                    change_df["_abs_shift"] = change_df["Rank Shift"].abs()
                    change_df = change_df.sort_values(
                        by=["_abs_shift", "Request"], ascending=[False, True]
                    ).drop(columns=["_abs_shift"])
                    render_theme_table(change_df, height_px=400)


# ============================================================================
# EXPORT TOOLS
# ============================================================================

st.header("Export Tools")


def format_export_datetime(value: Any) -> Optional[str]:
    """Keep exported timestamps readable, sortable, and importable."""
    if value in (None, "", "-"):
        return None
    if isinstance(value, datetime):
        dt = value
    else:
        try:
            dt = datetime.fromisoformat(str(value))
        except Exception:
            return str(value)
    return dt.strftime("%Y-%m-%d %H:%M")


def build_request_dataset_xlsx(df: pd.DataFrame) -> bytes:
    """Create an Excel version with readable column widths."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        raise RuntimeError("Excel export requires openpyxl. Install it with: pip install openpyxl") from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Request Dataset"

    headers = list(df.columns)
    worksheet.append(headers)
    for _, row in df.fillna("").iterrows():
        worksheet.append([row.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="EDE7F6")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    width_by_column = {
        "request_id": 14,
        "source": 12,
        "college": 10,
        "document_type": 46,
        "requester_type": 22,
        "urgency": 10,
        "submission_time": 20,
        "requirements_incomplete": 24,
        "requirements_partial": 22,
        "requirements_complete": 24,
        "payment_unpaid": 20,
        "payment_paid": 20,
        "ready_time": 20,
        "assignment_time": 20,
        "completion_time": 20,
    }
    for col_index, header in enumerate(headers, start=1):
        width = width_by_column.get(header, min(max(len(str(header)) + 4, 12), 28))
        worksheet.column_dimensions[get_column_letter(col_index)].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


request_dataset_rows = []
for item in results.get("generated_requests", []):
    submission_time = item.get("submission_time")
    requirements_partial_time = item.get("requirements_partial_time")
    requirements_complete_time = item.get("requirements_complete_time")
    payment_time = item.get("payment_time")
    request_dataset_rows.append(
        {
            "request_id": item.get("request_id"),
            "source": "custom" if item.get("is_custom") else "generated",
            "college": item.get("college"),
            "document_type": item.get("document_type"),
            "requester_type": item.get("requester_type"),
            "urgency": item.get("urgency"),
            "submission_time": format_export_datetime(submission_time),

            # Requirement lifecycle as event timestamps.
            "requirements_incomplete": format_export_datetime(submission_time),
            "requirements_partial": format_export_datetime(requirements_partial_time),
            "requirements_complete": format_export_datetime(requirements_complete_time),

            # Payment lifecycle as event timestamps.
            "payment_unpaid": format_export_datetime(submission_time) if payment_time else None,
            "payment_paid": format_export_datetime(payment_time),

            "ready_time": format_export_datetime(item.get("ready_time")),
            "assignment_time": format_export_datetime(item.get("assignment_time")),
            "completion_time": format_export_datetime(item.get("completion_time")),
        }
    )

if request_dataset_rows:
    request_dataset_df = pd.DataFrame(request_dataset_rows)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    st.download_button(
        "Download Request Dataset CSV",
        data=request_dataset_df.to_csv(index=False),
        file_name=f"request_dataset_{timestamp}.csv",
        mime="text/csv",
        use_container_width=True,
    )
    try:
        st.download_button(
            "Download Request Dataset Excel",
            data=build_request_dataset_xlsx(request_dataset_df),
            file_name=f"request_dataset_{timestamp}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    except RuntimeError as exc:
        st.caption(str(exc))

if engine.completed:
    export_df = pd.DataFrame(
        [
            {
                "request_id": req.request_id,
                "college": req.college,
                "document_type": req.document_type,
                "requester_status": getattr(req, "requester_type", "-"),
                "completeness_of_requirements": round(
                    float(getattr(req, "completeness_of_requirements", 0.0)),
                    4,
                ),
                "payment_status": getattr(req, "payment_status", "-"),
                "submission_time": format_export_datetime(req.submission_time),
                "assignment_time": format_export_datetime(req.assignment_time),
                "completion_time": format_export_datetime(req.completion_time),
                "queue_wait_hours": round((req.get_waiting_time_minutes() or 0.0) / 60.0, 4),
                "turnaround_days": round(req.get_turnaround_time_minutes() / 1440.0, 4),
                "assigned_staff": req.assigned_staff,
            }
            for req in engine.completed
        ]
    )

    csv_data = export_df.to_csv(index=False)
    st.download_button(
        "Download Results CSV",
        data=csv_data,
        file_name=f"simulation_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        mime="text/csv",
        use_container_width=True,
    )

if st.session_state.last_run_config:
    config_json = {
        "generated_at": datetime.now().isoformat(),
        "scheduler_type": results.get("scheduler_type"),
        "allocator_type": results.get("allocator_type"),
        "mode": "custom_sliders",
        "work_hours": results.get("work_hours"),
        "priority_weights": results.get("priority_weights"),
        "run_config": st.session_state.last_run_config.get("run_config", {}),
        "ui_config": st.session_state.last_run_config.get("ui_config", {}),
    }
    st.download_button(
        "Download Run Config JSON",
        data=json.dumps(config_json, indent=2),
        file_name=f"simulation_config_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
        mime="application/json",
        use_container_width=True,
    )
